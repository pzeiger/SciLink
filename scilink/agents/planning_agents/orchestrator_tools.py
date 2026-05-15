"""
Tool definitions and schemas for the PlanningOrchestratorAgent.
Supports both Google Gemini (function objects) and OpenAI (JSON schemas).
"""

from datetime import datetime
import json
import logging
import re
import pandas as pd
from pathlib import Path
from typing import Dict, Any, Callable, List, Optional
import hashlib


def _natural_sort_key(s):
    """Sort key that handles embedded numbers naturally (e.g., run_2 before run_10)."""
    return [int(c) if c.isdigit() else c.lower() for c in re.split(r'(\d+)', str(s))]

from .parser_utils import write_experiments_to_disk
from .instruct import (
    BO_OBJECTIVE_DISTILL_PROMPT,
    KNOWLEDGE_QUERY_CODEGEN_PROMPT,
    KNOWLEDGE_QUERY_DIRECTORY_CODEGEN_PROMPT,
)
from ..lit_agents.optimize_query import optimize_search_query, is_molecule_design_objective
from ...skills.loader import list_skills, load_skill


def _build_planning_skill_description(custom_skills: dict = None) -> str:
    """Build the ``skill`` parameter description for ``generate_initial_plan``.

    Auto-discovers built-in planning skill bundles (and any custom or
    graduated skills) so the orchestrator LLM can see which skills are
    available by name instead of having to be told. Mirrors the analyze-mode
    ``_build_skill_description`` helper, scoped to ``domain="planning"``.
    """
    parts = [
        "Optional domain skill: a built-in planning skill name or a path to "
        "a custom .md skill file. When set, the skill's validated domain "
        "rules are injected as mandatory constraints on the generated plan."
    ]

    try:
        names = list_skills(domain="planning")
    except Exception:
        names = []

    skill_descs = []
    for name in names:
        try:
            parsed = load_skill(name, domain="planning")
            desc = (parsed.get("meta") or {}).get("description")
            if not desc:
                desc = parsed.get("overview", "").split("\n")[0].strip()
            # Trim trailing punctuation so the join below stays clean.
            desc = desc.rstrip(".;,") if desc else desc
            skill_descs.append(f"'{name}' — {desc}" if desc else f"'{name}'")
        except Exception:
            skill_descs.append(f"'{name}'")
    if skill_descs:
        parts.append(f"Built-in planning skills: {'; '.join(skill_descs)}.")

    if custom_skills:
        parts.append(f"Custom skills: {sorted(custom_skills.keys())}.")

    return " ".join(parts)


class OrchestratorTools:
    """
    Manages tool definitions, schemas, and execution for the OrchestratorAgent.
    """
    
    def __init__(self, orchestrator_instance):
        """
        Args:
            orchestrator_instance: Reference to the parent OrchestratorAgent
        """
        self.orch = orchestrator_instance
        
        # Build function map and schemas
        self.functions_map: Dict[str, Callable] = {}
        self.openai_schemas: list = []
        self.gemini_functions: list = []
        
        self._register_all_tools()

    def _get_human_feedback_enabled(self) -> bool:
        """
        Get current human feedback setting from orchestrator.
        Returns True if not set (backwards compatible default).
        """
        return getattr(self.orch, '_enable_human_feedback', True)

    def _decode_categorical_recs(self, recs: Any, level_maps: Dict[str, List[str]]) -> Any:
        """Map integer-encoded categorical values back to their level names.

        Recommendations may be either a dict (single experiment) or a list of
        dicts (batch). Continuous values are passed through unchanged. The
        decoded value lookup uses the nearest integer index, which the
        MixedSingleTaskGP path already constrains to valid levels but other
        surrogates may return as a float.
        """
        if isinstance(recs, list):
            return [self._decode_categorical_recs(r, level_maps) for r in recs]
        if not isinstance(recs, dict):
            return recs
        out = dict(recs)
        for col, levels in level_maps.items():
            if col not in out:
                continue
            try:
                idx = int(round(float(out[col])))
            except (TypeError, ValueError):
                continue
            idx = max(0, min(idx, len(levels) - 1))
            out[col] = levels[idx]
        return out

    def _capture_input_types(self, column_roles: Dict, input_columns: List[str]) -> None:
        """Persist scalarizer input_types onto the orchestrator state.

        Filters to declared input columns only; missing entries default to
        "continuous" downstream. No-op when column_roles has no input_types
        field (backward-compat with older scalarizer outputs).
        """
        if not input_columns:
            return
        types_in = (column_roles or {}).get("input_types") or {}
        if not types_in:
            return
        filtered = {c: types_in[c] for c in input_columns if c in types_in}
        if filtered:
            self.orch.expected_input_types = filtered

    def _compute_file_hash(self, file_path: str) -> str:
        """Compute MD5 hash of file content for deduplication."""
        hasher = hashlib.md5()
        try:
            with open(file_path, 'rb') as f:
                for chunk in iter(lambda: f.read(8192), b''):
                    hasher.update(chunk)
            return hasher.hexdigest()
        except Exception as e:
            logging.warning(f"Could not compute hash for {file_path}: {e}")
            return ""


    def _distill_objective_for_bo(self, target_cols: list) -> str:
        """
        Distill a verbose user objective into a concise BO-relevant objective.
        Uses the orchestrator's LLM to extract only optimization targets and
        directions. Result is cached on self.orch._distilled_objective.
        """
        raw = self.orch.objective
        # Skip distillation for short/default objectives
        if (not raw
                or raw == "Undefined Research Goal"
                or len(raw) <= 200):
            return raw

        cached = getattr(self.orch, '_distilled_objective', None)
        if cached is not None:
            return cached

        try:
            prompt = BO_OBJECTIVE_DISTILL_PROMPT.format(
                objective=raw,
                target_cols=", ".join(target_cols),
            )
            resp = self.orch.bo.model.generate_content(
                [prompt], generation_config=self.orch.bo.generation_config
            )
            distilled = resp.text.strip()
            if distilled:
                print(f"    🎯 Distilled objective: {distilled}")
                self.orch._distilled_objective = distilled
                return distilled
        except Exception as e:
            logging.warning(f"Objective distillation failed, using original: {e}")

        return raw

    def _parse_result_input(self, result_data: str):
        """
        Helper to parse result_data into appropriate format.

        Returns:
            - String (text input)
            - String (single file path)
            - List of strings (multiple file paths)
        """
        if len(result_data) < 500:  # Reasonable path length
            try:
                # Check if it's comma-separated file paths
                if ',' in result_data:
                    paths = [p.strip() for p in result_data.split(',')]
                    valid_paths = []
                    for p in paths:
                        resolved, error = self._resolve_data_path(p)
                        if not error:
                            valid_paths.append(resolved)

                    if valid_paths:
                        print(f"    (Detected {len(valid_paths)} file paths)")
                        return valid_paths

                # Check if it's a single file path (try session-aware resolution)
                resolved, error = self._resolve_data_path(result_data.strip())
                if not error and Path(resolved).is_file():
                    print(f"    (Detected file path: {Path(resolved).name})")
                    return str(resolved)

                # Not a valid path - treat as text
                text_preview = result_data[:100] + "..." if len(result_data) > 100 else result_data
                print(f"    (Processing text input: '{text_preview}')")
                return result_data

            except (OSError, ValueError, RuntimeError):
                # Not a valid path - treat as text
                text_preview = result_data[:100] + "..." if len(result_data) > 100 else result_data
                print(f"    (Processing text input: '{text_preview}')")
                return result_data
        else:
            # Too long to be a path - treat as text
            text_preview = result_data[:100] + "..." if len(result_data) > 100 else result_data
            print(f"    (Processing text input: '{text_preview}')")
            return result_data
        
    def _collect_scalarizer_context(self, payload) -> list:
        """Collect scalarizer metrics and plot for files that were already analyzed.

        Checks whether any file path in *payload* has a corresponding entry in
        ``self.orch.analyzed_files``.  If so, appends:
        - The computed metrics from ``optimization_data.csv`` as a text summary
        - The scalarizer debug plot image path (if it exists)

        Returns a list of extra items to append to the refinement payload
        (may be empty).
        """
        paths = payload if isinstance(payload, list) else [payload]
        extras = []
        seen_bo = False

        for item in paths:
            if not isinstance(item, str):
                continue

            abs_path = str(Path(item).resolve())
            if abs_path not in self.orch.analyzed_files:
                continue

            # Append computed metrics from optimization_data.csv (once)
            if not seen_bo and self.orch.bo_data_path.exists():
                try:
                    df = pd.read_csv(self.orch.bo_data_path)
                    summary = (
                        f"SCALARIZER ANALYSIS RESULTS (computed metrics):\n"
                        f"{df.to_string(index=False)}"
                    )
                    extras.append(summary)
                    seen_bo = True
                    print(f"    📊 Attached scalarizer metrics ({len(df)} rows)")
                except Exception:
                    pass

            # Append debug plot if it exists
            stem = Path(item).stem
            plot_path = self.orch.base_dir / "scalarizer_outputs" / f"debug_{stem}.png"
            if plot_path.exists():
                extras.append(str(plot_path))
                print(f"    📈 Attached scalarizer plot: {plot_path.name}")

        return extras

    def _resolve_knowledge_paths(self, knowledge_paths: str | None) -> list[str] | None:
        """Resolve knowledge paths with fallback to orchestrator's knowledge_dir.

        If the LLM provides explicit paths, use those.  Otherwise fall back to
        ``self.orch.knowledge_dir`` so the KB can match sources from previous
        sessions (stable path) instead of rebuilding from a session-specific dir.
        """
        if knowledge_paths:
            paths = [p.strip() for p in knowledge_paths.split(",") if p.strip()]
            if paths:
                return paths
        # Fallback: use orchestrator's configured knowledge directory
        if self.orch.knowledge_dir and self.orch.knowledge_dir.exists():
            return [str(self.orch.knowledge_dir)]
        return None

    @staticmethod
    def _build_objective_guidance(n_data: int, numeric_cols: list) -> dict:
        """Return data-aware guidance on how many targets the data can support."""
        # Estimate max feasible inputs (assume at least 2)
        n_numeric = len(numeric_cols)
        est_inputs = max(2, n_numeric // 2)

        supported = {}
        for n_t in range(1, min(n_numeric, 5)):
            needed = 5 * est_inputs * n_t
            supported[n_t] = {"min_recommended": needed, "feasible": n_data >= needed}

        max_feasible = max((k for k, v in supported.items() if v["feasible"]), default=1)

        return {
            "data_points": n_data,
            "numeric_columns": n_numeric,
            "supported_targets": supported,
            "max_feasible_targets": max_feasible,
            "recommendation": (
                f"With {n_data} data points, up to {max_feasible} target(s) can be "
                f"optimized reliably. Pick the {max_feasible} most important target(s) "
                f"aligned with the stated objective. Additional targets can be added "
                f"later when more data is collected."
            ),
        }

    def _resolve_data_path(self, path_input: str) -> tuple[str, str]:
        """
        Resolves user input to actual file path with fuzzy matching for typos.
        
        Returns:
            (resolved_path, None) on success
            (None, error_json) on failure (with suggestions if available)
        """
        from difflib import get_close_matches
        
        path = Path(path_input.strip())
        
        # Case 1: Path exists as-is
        if path.exists():
            return str(path), None
        
        # Case 2: Try common extensions if no extension provided
        if not path.suffix:
            for ext in ['.csv', '.xlsx', '.xls']:
                candidate = path.with_suffix(ext)
                if candidate.exists():
                    print(f"    🔍 Resolved: {path.name} → {candidate.name}")
                    return str(candidate), None
        
        # Case 3: Try in common data folders (session dirs first, then cwd-relative)
        session = self.orch.base_dir
        search_folders = [
            str(session / "uploads"),
            str(session / "uploads" / "series"),
            str(session / "data"),
            str(session),
            './experimental_results', './data', './results', './',
        ]
        all_candidates = []  # Track all files we find for fuzzy matching
        
        if not path.is_absolute():
            stem = path.stem if path.suffix else path.name
            
            for folder in search_folders:
                folder_path = Path(folder)
                if not folder_path.exists():
                    continue
                
                # Collect all data files in this folder
                for ext in ['.csv', '.xlsx', '.xls']:
                    all_candidates.extend(folder_path.glob(f"*{ext}"))
                
                # Try exact match with provided extension
                if path.suffix:
                    candidate = folder_path / path.name
                    if candidate.exists():
                        print(f"    🔍 Found: {path.name} in {folder}/")
                        return str(candidate), None
                
                # Try common extensions
                for ext in ['.csv', '.xlsx', '.xls']:
                    candidate = folder_path / f"{stem}{ext}"
                    if candidate.exists():
                        print(f"    🔍 Found: {stem}{ext} in {folder}/")
                        return str(candidate), None
        
        # Case 4: File not found - use fuzzy matching to suggest alternatives
        if all_candidates:
            # Get filenames without path
            candidate_names = [f.name for f in all_candidates]
            
            # Try fuzzy match on the input filename
            input_name = path.name
            matches = get_close_matches(input_name, candidate_names, n=3, cutoff=0.6)
            
            if matches:
                # Find full paths for the matches
                suggested_files = []
                for match in matches:
                    for candidate in all_candidates:
                        if candidate.name == match:
                            suggested_files.append(str(candidate))
                            break
                
                return None, json.dumps({
                    "status": "error",
                    "message": f"File not found: {path_input}",
                    "did_you_mean": matches,
                    "full_paths": suggested_files,
                    "hint": f"Did you mean '{matches[0]}'? Use: primary_data_set='{suggested_files[0]}'"
                })
        
        # No matches found at all
        return None, json.dumps({
            "status": "error",
            "message": f"Could not find file: {path_input}",
            "searched_in": [str(f) for f in search_folders if Path(f).exists()],
            "hint": "Check filename spelling or use /files command to see available files"
        })
    
    def _register_all_tools(self):
        """Register all tools with both OpenAI and Gemini formats."""
        
        # 0. LIST WORKSPACE FILES
        def list_workspace_files():
            """Lists files in the campaign directory including analysis artifacts."""
            print(f"  ⚡ Tool: Listing files in {self.orch.base_dir}...")
            files = [f.name for f in self.orch.base_dir.iterdir() if f.is_file()]
            artifacts_dir = self.orch.base_dir / "analysis_artifacts"
            artifact_names = []
            if artifacts_dir.exists():
                 artifact_names = [f"analysis_artifacts/{f.name}" for f in artifacts_dir.iterdir() if f.is_file()]
            
            all_files = files + artifact_names
            
            # Include data point count for optimization readiness
            data_count = 0
            if self.orch.bo_data_path.exists():
                try:
                    df = pd.read_csv(self.orch.bo_data_path)
                    data_count = len(df)
                except:
                    pass
            
            return json.dumps({
                "status": "success",
                "files": all_files,
                "data_points_collected": data_count,
                "optimization_ready": data_count >= 3,
                "active_analysis_script": Path(self.orch.active_scalarizer_script).name if self.orch.active_scalarizer_script else None
            })
        
        self._register_tool(
            func=list_workspace_files,
            name="list_workspace_files",
            description="Lists files in the session directory (checkpoints, analysis artifacts, etc.). User data files may exist outside the session folder.",
            parameters={}
        )
        
        # --- LITERATURE SEARCH TOOL ---
        def search_literature(objective: str, search_type: str = "hypothesis_context"):
            """
            Searches scientific literature using the FutureHouse Edison API.
            Call this BEFORE generate_initial_plan to enrich the plan with
            external literature context.
            """
            if not self.orch.lit_agent:
                return json.dumps({
                    "status": "error",
                    "message": "Literature search not available (no FutureHouse API key configured)"
                })

            valid_types = ("hypothesis_context", "economic_data", "fitting_models")
            if search_type not in valid_types:
                return json.dumps({
                    "status": "error",
                    "message": f"Invalid search_type '{search_type}'. Must be one of: {', '.join(valid_types)}"
                })

            print(f"  ⚡ Tool: Searching literature ({search_type}) for '{objective[:80]}...'")

            try:
                clean_query = optimize_search_query(
                    objective=objective, model=self.orch.planner.model
                )

                search_methods = {
                    "hypothesis_context": self.orch.lit_agent.search_for_hypothesis_context,
                    "economic_data": self.orch.lit_agent.search_for_economic_data,
                    "fitting_models": self.orch.lit_agent.search_for_fitting_models,
                }
                lit_res = search_methods[search_type](clean_query)

                if lit_res['status'] != 'success':
                    return json.dumps({
                        "status": lit_res['status'],
                        "message": lit_res.get('message', 'Literature search did not succeed')
                    })

                # Save to file (distinct per search_type to avoid overwrites)
                lit_path = self.orch.base_dir / f"literature_search_{search_type}.md"
                with open(lit_path, 'w') as f:
                    f.write(f"# Literature Search Results ({search_type})\n\n")
                    f.write(lit_res['content'])

                print(f"  ✅ Literature search completed. Saved to {lit_path.name}")

                return json.dumps({
                    "status": "success",
                    "file_path": str(lit_path),
                    "content_preview": lit_res['content'][:500] + "..." if len(lit_res['content']) > 500 else lit_res['content'],
                    "hint": "Pass file_path as literature_context to generate_initial_plan()"
                })

            except Exception as e:
                logging.error(f"Literature search error: {e}", exc_info=True)
                return json.dumps({"status": "error", "message": str(e)})

        self._register_tool(
            func=search_literature,
            name="search_literature",
            description=(
                "Searches scientific literature via FutureHouse Edison API. "
                "Call BEFORE generate_initial_plan() to enrich the plan with external context. "
                "Pass the returned file_path as literature_context to generate_initial_plan()."
            ),
            parameters={
                "objective": {"type": "string", "description": "Research objective or question to search for"},
                "search_type": {
                    "type": "string",
                    "description": "Type of search: 'hypothesis_context' (default, for planning), 'economic_data' (for TEA), or 'fitting_models' (for curve fitting)",
                    "enum": ["hypothesis_context", "economic_data", "fitting_models"]
                }
            },
            required=["objective"]
        )

        # --- MOLECULES QUERY TOOL ---
        def query_molecules(objective: str):
            """
            Queries the FutureHouse Molecules agent for molecular design,
            synthesis planning, or cheminformatics tasks.
            Call this BEFORE generate_initial_plan() when the objective
            involves molecular design or discovery.
            """
            if not self.orch.mol_agent:
                return json.dumps({
                    "status": "error",
                    "message": "Molecules agent not available (no FutureHouse API key configured)"
                })

            # Guard: only proceed for genuine molecule design objectives
            if not is_molecule_design_objective(objective, self.orch.planner.model):
                return json.dumps({
                    "status": "skipped",
                    "message": "Objective does not appear to involve molecular design or synthesis planning. Skipping molecules query."
                })

            print(f"  ⚡ Tool: Querying MOLECULES agent for '{objective[:80]}...'")

            try:
                mol_res = self.orch.mol_agent.query(objective)

                if mol_res['status'] != 'success':
                    return json.dumps({
                        "status": mol_res['status'],
                        "message": mol_res.get('message', 'Molecules query did not succeed')
                    })

                # Save to file
                mol_path = self.orch.base_dir / "molecule_design.md"
                with open(mol_path, 'w') as f:
                    f.write("# Molecular Design & Synthesis Planning Results\n\n")
                    f.write(mol_res['content'])

                print(f"  ✅ Molecules query completed. Saved to {mol_path.name}")

                return json.dumps({
                    "status": "success",
                    "file_path": str(mol_path),
                    "content_preview": mol_res['content'][:500] + "..." if len(mol_res['content']) > 500 else mol_res['content'],
                    "hint": "Pass file_path as molecule_context to generate_initial_plan()"
                })

            except Exception as e:
                logging.error(f"Molecules query error: {e}", exc_info=True)
                return json.dumps({"status": "error", "message": str(e)})

        self._register_tool(
            func=query_molecules,
            name="query_molecules",
            description=(
                "Queries the FutureHouse Molecules agent for molecular design, synthesis planning, "
                "or cheminformatics. Call BEFORE generate_initial_plan() when the objective involves "
                "molecule design or discovery. Pass the returned file_path as molecule_context."
            ),
            parameters={
                "objective": {"type": "string", "description": "Molecular design or synthesis objective"}
            },
            required=["objective"]
        )

        # 1. GENERATE INITIAL PLAN
        def generate_initial_plan(
            specific_objective: str = None,
            knowledge_paths: str = None,
            primary_data_set: str = None,
            additional_context: str = None,
            skill: str = None,
            literature_context: str = None,
            molecule_context: str = None
        ):
            """
            Generates experimental plan (science strategy only, no code).

            Note: code_paths parameter is deprecated. Use generate_implementation_code()
            as a separate step to add code after plan approval.
            """
            obj = specific_objective if specific_objective else self.orch.objective
            print(f"  ⚡ Tool: Generating Initial Plan for '{obj}'...")

            # Resolve knowledge paths (with fallback to orchestrator dir)
            knowledge_list = self._resolve_knowledge_paths(knowledge_paths)
            if knowledge_list:
                # Validate paths
                invalid_paths = [p for p in knowledge_list if not Path(p).exists()]
                if invalid_paths:
                    return json.dumps({
                        "status": "error",
                        "message": f"Knowledge paths not found: {', '.join(invalid_paths)}",
                        "hint": "Check folder names and spelling"
                    })
                print(f"    📚 Knowledge sources: {knowledge_list}")

            # Parse primary dataset - UPDATED LOGIC
            primary_dataset = None
            if primary_data_set:
                # Try to resolve the path
                resolved_path, error = self._resolve_data_path(primary_data_set)
                
                if error:
                    return error  # Return the error JSON with suggestions
                
                path = Path(resolved_path)
                
                # Now handle resolved path
                if path.is_file():
                    primary_dataset = {"file_path": str(path)}
                    print(f"    📊 Primary data: {path.name}")
                    
                elif path.is_dir():
                    # Directory - check how many data files
                    all_files = []
                    for ext in ['*.csv', '*.xlsx', '*.xls']:
                        all_files.extend(path.glob(ext))
                    
                    if not all_files:
                        return json.dumps({
                            "status": "error",
                            "message": f"No data files (.csv, .xlsx, .xls) found in: {primary_data_set}",
                            "hint": "Add data files to the folder or specify a different path"
                        })
                    
                    elif len(all_files) == 1:
                        # Only one file - use it automatically
                        primary_dataset = {"file_path": str(all_files[0])}
                        print(f"    📊 Primary data (auto-selected): {all_files[0].name}")
                        
                    else:
                        # Multiple files - require user to specify
                        file_list = sorted([f.name for f in all_files], key=_natural_sort_key)
                        return json.dumps({
                            "status": "error",
                            "message": f"Multiple data files found in '{primary_data_set}'",
                            "available_files": file_list,
                            "file_count": len(file_list),
                            "hint": f"Please specify which file to use. Example: primary_data_set='./experimental_results/{file_list[0]}'"
                        })
            
            # Build context
            context_parts = []
            
            if additional_context:
                context_parts.append(f"User Requirements: {additional_context}")
                print(f"    ℹ️  User context: {additional_context[:60]}...")
            
            # Auto-include TEA results
            if self.orch.latest_tea_results:
                tea_summary = self.orch.latest_tea_results.get('summary', '')
                context_parts.append(f"Economic Analysis Results: {tea_summary}")
                print(f"    💰 Including TEA results in context")
            
            context_dict = None
            if context_parts:
                context_dict = {"user_context": "\n\n".join(context_parts)}
            
            # Resolve skill: use provided value or fall back to orchestrator's active skill
            effective_skill = skill or getattr(self.orch, '_active_skill', None)

            # Build external_context from literature/molecule files or raw text
            external_context_parts = []
            saved_extras = []
            if literature_context:
                lp = Path(literature_context)
                if lp.is_file():
                    lit_text = lp.read_text()
                    external_context_parts.append(lit_text)
                    saved_extras.append(str(lp))
                    print(f"    📚 Literature context from: {lp.name}")
                else:
                    external_context_parts.append(literature_context)
            if molecule_context:
                mp = Path(molecule_context)
                if mp.is_file():
                    mol_text = mp.read_text()
                    external_context_parts.append(
                        "## Molecular Design & Synthesis Planning\n" + mol_text
                    )
                    saved_extras.append(str(mp))
                    print(f"    🧪 Molecule context from: {mp.name}")
                else:
                    external_context_parts.append(
                        "## Molecular Design & Synthesis Planning\n" + molecule_context
                    )

            ext_ctx = "\n\n".join(external_context_parts) if external_context_parts else None

            try:
                plan = self.orch.planner.generate_plan(
                    objective=obj,
                    knowledge_paths=knowledge_list,
                    primary_data_set=primary_dataset,
                    additional_context=context_dict,
                    enable_human_feedback=self._get_human_feedback_enabled(),
                    reset_state=False,
                    skill=effective_skill,
                    external_context=ext_ctx
                )

                # Store skill on orchestrator for downstream tools
                if effective_skill:
                    self.orch._active_skill = effective_skill

                if plan.get("error"):
                    return json.dumps({
                        "status": "error",
                        "message": plan.get("error")
                    })

                # Save
                output_path = self.orch.base_dir / "plan.json"
                with open(output_path, 'w') as f:
                    json.dump(plan, f, indent=2)

                # If literature came from the deprecated internal path, save it
                if not literature_context and plan.get("literature_search"):
                    lit_path = self.orch.base_dir / "literature_search.md"
                    with open(lit_path, 'w') as f:
                        f.write("# Literature Search Results\n\n")
                        f.write(plan["literature_search"])
                    saved_extras.append(str(lit_path))

                # Generate HTML
                from .html_generator import HTMLReportGenerator
                html_path = self.orch.base_dir / "plan.html"
                generator = HTMLReportGenerator(self.orch.planner.state)
                generator.generate(str(html_path))

                num_experiments = len(plan.get('proposed_experiments', []))

                result = {
                    "status": "success",
                    "iteration": plan.get('iteration'),
                    "num_experiments": num_experiments,
                    "output_path": str(output_path),
                    "html_report": str(html_path),
                    "knowledge_used": knowledge_list is not None,
                    "primary_data_used": primary_dataset is not None,
                    "tea_context_included": self.orch.latest_tea_results is not None,
                    "hint": "Use generate_implementation_code() to add executable code"
                }
                if saved_extras:
                    result["external_results_files"] = saved_extras
                return json.dumps(result)
                
            except Exception as e:
                logging.error(f"Plan generation error: {e}", exc_info=True)
                return json.dumps({
                    "status": "error",
                    "message": str(e)
                })

        # Register it
        self._register_tool(
            func=generate_initial_plan,
            name="generate_initial_plan",
            description=(
                "Generates experimental plan (science strategy only, no implementation code). "
                "Automatically includes previous TEA results if available. "
                "Can use: papers/reports, experimental data, lab constraints."
            ),
            parameters={
                "specific_objective": {"type": "string", "description": "Research objective"},
                "knowledge_paths": {"type": "string", "description": "Comma-separated paths to papers/reports/docs folders"},
                "primary_data_set": {"type": "string", "description": "Path to experimental data file or folder"},
                "additional_context": {"type": "string", "description": "Lab constraints, equipment, reagents, budget, etc."},
                "skill": {
                    "type": "string",
                    "description": _build_planning_skill_description(
                        getattr(self.orch, "_custom_skills", None)
                    ),
                },
                "literature_context": {"type": "string", "description": "File path or text from search_literature() tool. Provides external scientific literature context."},
                "molecule_context": {"type": "string", "description": "File path or text from query_molecules() tool. Provides molecular design / synthesis context."}
            },
            required=[]
        )

        # 2. GENERATE IMPLEMENTATION CODE
        def generate_implementation_code(code_paths: str = None):
            """
            Adds implementation code to the most recent experimental plan.
            Use after generate_initial_plan() to map experiments to executable code.
            
            Args:
                code_paths: Comma-separated paths to code folders. 
                        Optional if Code KB already loaded at startup.
            """
            
            if not self.orch.planner.state or not self.orch.planner.state.get("current_plan"):
                return json.dumps({
                    "status": "error",
                    "message": "No active plan. Generate a plan first using generate_initial_plan()"
                })
            
            current_plan = self.orch.planner.state["current_plan"]
            
            # Check if already has code
            if current_plan.get("proposed_experiments"):
                has_code = any(exp.get("implementation_code") for exp in current_plan["proposed_experiments"])
                if has_code:
                    return json.dumps({
                        "status": "warning",
                        "message": "Plan already has implementation code",
                        "hint": "Generate a new plan if you want to change the code source"
                    })
            
            print(f"  ⚡ Tool: Generating implementation code for existing plan...")

            kb_available = (self.orch.planner.kb_code.index and 
                            self.orch.planner.kb_code.index.ntotal > 0)

            if not kb_available and not code_paths:
                return json.dumps({
                    "status": "error",
                    "message": "No Code Knowledge Base available",
                    "hint": "Provide code_paths parameter (e.g., code_paths='./opentrons_api,./automation_lib')",
                    "available_options": [
                        "Option 1: Specify code_paths='./your_code_folder'",
                        "Option 2: If code exists, check folder name and path"
                    ]
                })
            
            # Parse code paths
            code_list = []
            if code_paths:
                code_list = [p.strip() for p in code_paths.split(',') if p.strip()]
                
                # Validate paths (only if code_paths was provided)
                invalid_paths = []
                for path in code_list:
                    if not Path(path).exists():
                        invalid_paths.append(path)
                
                if invalid_paths:
                    # Check for common typos
                    suggestions = []
                    for invalid in invalid_paths:
                        parent = Path(invalid).parent
                        if parent.exists():
                            similar = [f.name for f in parent.iterdir() 
                                    if f.is_dir() and invalid.lower() in f.name.lower()]
                            if similar:
                                suggestions.append(f"Did you mean './{similar[0]}'?")
                    
                    hint = "Check folder names and spelling."
                    if suggestions:
                        hint += " " + " ".join(suggestions)
                    
                    return json.dumps({
                        "status": "error",
                        "message": f"Code paths not found: {', '.join(invalid_paths)}",
                        "hint": hint
                    })
                
                print(f"    💻 Code sources: {code_list}")
            elif kb_available:
                print(f"    💻 Using existing Code KB ({self.orch.planner.kb_code.index.ntotal} vectors)")
            
            try:
                updated_plan = self.orch.planner.generate_implementation_code(
                    plan=current_plan,
                    code_paths=code_list,
                    enable_human_feedback=self._get_human_feedback_enabled()
                )
                
                if updated_plan.get("error"):
                    return json.dumps({
                        "status": "error",
                        "message": updated_plan.get("error")
                    })
                
                # Save
                output_path = self.orch.base_dir / "plan.json"
                with open(output_path, 'w') as f:
                    json.dump(updated_plan, f, indent=2)
                
                # Regenerate HTML
                from .html_generator import HTMLReportGenerator
                html_path = self.orch.base_dir / "plan.html"
                generator = HTMLReportGenerator(self.orch.planner.state)
                generator.generate(str(html_path))
                
                # Check if any experiments actually got code
                experiments = updated_plan.get("proposed_experiments", [])
                has_code = any(
                    exp.get("implementation_code")
                    for exp in experiments
                )

                if not has_code:
                    return json.dumps({
                        "status": "error",
                        "message": "Code generation failed — no executable code was produced for any experiment.",
                        "hint": "This may be due to an LLM API timeout or error. Try again.",
                        "output_path": str(output_path),
                        "html_report": str(html_path)
                    })

                # Save scripts to output folder
                final_out = str(self.orch.base_dir / "output_scripts")
                print(f"\n--- Saving Scripts to: {final_out} ---")
                write_experiments_to_disk(updated_plan, final_out)

                return json.dumps({
                    "status": "success",
                    "message": "Implementation code added to plan",
                    "output_path": str(output_path),
                    "html_report": str(html_path),
                    "scripts_saved_to": final_out,
                    "code_sources_used": code_list
                })
                
            except Exception as e:
                logging.error(f"Code generation error: {e}", exc_info=True)
                return json.dumps({
                    "status": "error",
                    "message": str(e)
                })

        # Register it
        self._register_tool(
            func=generate_implementation_code,
            name="generate_implementation_code",
            description=(
                "Generates executable implementation code for the most recent experimental plan. "
                "Maps experimental steps to code using API documentation and example repositories. "
                "Use after generate_initial_plan() once the scientific strategy is approved. "
                "If Code KB already loaded, code_paths is optional."
            ),
            parameters={
                "code_paths": {
                    "type": "string",
                    "description": (
                        "Comma-separated paths to SOURCE CODE or API documentation folders "
                        "(e.g., './opentrons_api,./automation_lib'). "
                        "Must contain .py, .js, or other code files — NOT scientific papers, "
                        "PDFs, or literature. Do NOT pass the knowledge directory here. "
                        "OPTIONAL if Code Knowledge Base is already loaded. "
                        "REQUIRED if no Code KB exists."
                    )
                }
            },
            required=[]
        )
        
        # 3. RUN ECONOMIC ANALYSIS
        def run_economic_analysis(
            focus_topic: str = None,
            knowledge_paths: str = None,
            primary_data_set: str = None,
            additional_context: str = None,
            literature_context: str = None
        ):
            """Performs Technoeconomic Analysis (TEA)."""
            obj = focus_topic if focus_topic else self.orch.objective
            print(f"  ⚡ Tool: Running TEA for '{obj}'...")

            # Resolve knowledge paths (with fallback to orchestrator dir)
            knowledge_list = self._resolve_knowledge_paths(knowledge_paths)
            if knowledge_list:
                print(f"    📚 Knowledge sources: {knowledge_list}")

            # Parse primary dataset
            primary_dataset = None
            if primary_data_set:
                # Try to resolve the path
                resolved_path, error = self._resolve_data_path(primary_data_set)
                
                if error:
                    return error  # Return the error JSON with suggestions
                
                path = Path(resolved_path)
                
                # Now handle resolved path
                if path.is_file():
                    primary_dataset = {"file_path": str(path)}
                    print(f"    📊 Primary data: {path.name}")
                    
                elif path.is_dir():
                    # Directory - check how many data files
                    all_files = []
                    for ext in ['*.csv', '*.xlsx', '*.xls']:
                        all_files.extend(path.glob(ext))
                    
                    if not all_files:
                        return json.dumps({
                            "status": "error",
                            "message": f"No data files (.csv, .xlsx, .xls) found in: {primary_data_set}",
                            "hint": "Add data files to the folder or specify a different path"
                        })
                    
                    elif len(all_files) == 1:
                        # Only one file - use it automatically
                        primary_dataset = {"file_path": str(all_files[0])}
                        print(f"    📊 Primary data (auto-selected): {all_files[0].name}")
                        
                    else:
                        # Multiple files - require user to specify
                        file_list = sorted([f.name for f in all_files], key=_natural_sort_key)
                        return json.dumps({
                            "status": "error",
                            "message": f"Multiple data files found in '{primary_data_set}'",
                            "available_files": file_list,
                            "file_count": len(file_list),
                            "hint": f"Please specify which file to use. Example: primary_data_set='./experimental_results/{file_list[0]}'"
                        })
            
            try:
                # Resolve literature context
                ext_ctx = None
                if literature_context:
                    lp = Path(literature_context)
                    ext_ctx = lp.read_text() if lp.is_file() else literature_context
                    print(f"    📚 Literature context from: {lp.name if lp.is_file() else 'inline text'}")

                res = self.orch.planner.perform_technoeconomic_analysis(
                    objective=obj,
                    knowledge_paths=knowledge_list,
                    primary_data_set=primary_dataset,
                    output_json_path=str(self.orch.base_dir / "tea_analysis.json"),
                    external_context=ext_ctx
                )
                
                if res.get("error"):
                    return json.dumps({
                        "status": "error",
                        "message": res.get("error")
                    })
                
                summary = res.get('technoeconomic_assessment', {}).get('summary', 'No summary')
                
                # Store TEA results in orchestrator state
                self.orch.latest_tea_results = {
                    "summary": summary,
                    "full_analysis": res.get('technoeconomic_assessment'),
                    "timestamp": datetime.now().isoformat()
                }
                print(f"    ✅ TEA results stored for future planning")
                
                return json.dumps({
                    "status": "success",
                    "summary": summary,
                    "output_path": str(self.orch.base_dir / "tea_analysis.json"),
                    "html_report": str(self.orch.base_dir / "tea_analysis.html"),
                    "hint": "These results will automatically inform future generate_initial_plan calls"
                })
                
            except Exception as e:
                logging.error(f"TEA error: {e}", exc_info=True)
                return json.dumps({
                    "status": "error",
                    "message": str(e)
                })

        self._register_tool(
            func=run_economic_analysis,
            name="run_economic_analysis",
            description=(
                "Performs Technoeconomic Analysis (TEA) to assess economic viability, costs, market fit. "
                "Can incorporate papers and experimental data."
            ),
            parameters={
                "focus_topic": {
                    "type": "string",
                    "description": "Specific technology/process to analyze"
                },
                "knowledge_paths": {
                    "type": "string",
                    "description": "Comma-separated folder paths with papers/PDFs"
                },
                "primary_data_set": {
                    "type": "string",
                    "description": "Path to experimental data file or folder"
                },
                "additional_context": {
                    "type": "string",
                    "description": "Any other relevant context (constraints, requirements, etc.)"
                },
                "literature_context": {
                    "type": "string",
                    "description": "File path or text from search_literature(search_type='economic_data'). Provides external economic literature context."
                }
            },
            required=[]
        )

        # 4. REFINE PLAN (based on results)
        def refine_plan_with_results(
            result_data: str,
            use_literature_rag: bool = False,
            literature_context: str = None,
            molecule_context: str = None,
            additional_context: str = None
        ):
            """
            Refines the experimental plan (science strategy only) based on results.

            Use this for:
            - Strategic pivots or failures
            - Qualitative observations
            - Visual analysis of plots/images
            - When experiments didn't go as expected

            Supports multiple input formats:
            - Text: "Yield was 12%, precipitation observed"
            - File path: "./data.csv" or "./plot.png"
            - Comma-separated files: "./data.csv,./plot.png"
            """
            print(f"  ⚡ Tool: Refining Plan based on Results...")

            # Parse input - handle both single paths and comma-separated lists
            payload = self._parse_result_input(result_data)

            # Enrich with scalarizer metrics and plot if the file was already analyzed
            extras = self._collect_scalarizer_context(payload)
            if extras:
                if isinstance(payload, list):
                    payload.extend(extras)
                else:
                    payload = [payload] + extras

            # Build external context from literature/molecule files or raw text
            ext_parts = []
            if literature_context:
                lp = Path(literature_context)
                ext_parts.append(lp.read_text() if lp.is_file() else literature_context)
                print(f"    📚 Literature context provided")
            else:
                # Auto-load hypothesis context from session if available
                lit_path = self.orch.base_dir / "literature_search_hypothesis_context.md"
                if lit_path.is_file():
                    ext_parts.append(lit_path.read_text())
                    print(f"    📚 Auto-loaded literature hypothesis context from session")
            if molecule_context:
                mp = Path(molecule_context)
                mol_text = mp.read_text() if mp.is_file() else molecule_context
                ext_parts.append("## Molecular Design & Synthesis Planning\n" + mol_text)
                print(f"    🧪 Molecule context provided")
            if additional_context:
                ext_parts.append(f"## Additional Context\n{additional_context}")
                print(f"    ℹ️  Additional context provided")
            ext_ctx = "\n\n".join(ext_parts) if ext_parts else None

            try:
                plan = self.orch.planner.refine_plan(
                    results=payload,
                    enable_human_feedback=self._get_human_feedback_enabled(),
                    use_literature_rag=use_literature_rag,
                    external_context=ext_ctx
                )
                
                if plan.get("error"):
                    return json.dumps({
                        "status": "error",
                        "message": plan.get("error")
                    })
                
                # Save
                output_path = self.orch.base_dir / "plan.json"
                with open(output_path, 'w') as f:
                    json.dump(plan, f, indent=2)

                # Generate HTML
                from .html_generator import HTMLReportGenerator
                html_path = self.orch.base_dir / "plan.html"
                generator = HTMLReportGenerator(self.orch.planner.state)
                generator.generate(str(html_path))

                return json.dumps({
                    "status": "success",
                    "iteration": plan.get('iteration'),
                    "num_experiments": len(plan.get('proposed_experiments', [])),
                    "output_path": str(output_path),
                    "html_report": str(html_path),
                    "hint": "Use refine_implementation_code() to update executable code"
                })
                
            except Exception as e:
                logging.error(f"Plan refinement error: {e}", exc_info=True)
                return json.dumps({
                    "status": "error",
                    "message": str(e)
                })
        
        self._register_tool(
            func=refine_plan_with_results,
            name="refine_plan_with_results",
            description=(
                "Refines experimental plan (science strategy only) based on results. "
                "Handles text descriptions, single file paths, or comma-separated files. "
                "Use for: failures, pivots, qualitative observations, or visual analysis. "
                "Does NOT update implementation code - use refine_implementation_code() for that."
            ),
            parameters={
                "result_data": {
                    "type": "string",
                    "description": "Experimental results (text, file path, or comma-separated files)"
                },
                "use_literature_rag": {
                    "type": "boolean",
                    "description": "Search local knowledge base for relevant context. Default: false."
                },
                "literature_context": {
                    "type": "string",
                    "description": "File path or text from search_literature() tool. Provides external scientific literature context for refinement."
                },
                "molecule_context": {
                    "type": "string",
                    "description": "File path or text from query_molecules() tool. Provides molecular design / synthesis context for refinement."
                },
                "additional_context": {
                    "type": "string",
                    "description": "Extra context (e.g., reference data from query_knowledge_data, constraints, observations) to inform refinement."
                }
            },
            required=["result_data"]
        )
        
        # 4b. ADJUST PLAN FOR CONSTRAINTS (pre-execution)
        def adjust_plan_for_constraints(constraint_description: str):
            """
            Adjusts the experimental plan for implementation or instrument
            constraints discovered during protocol/code generation.
            Does NOT increment the iteration — the experiment hasn't run yet.
            """
            print(f"  ⚡ Tool: Adjusting plan for implementation constraints...")

            try:
                plan = self.orch.planner.adjust_plan_for_constraints(
                    constraint_description=constraint_description,
                    enable_human_feedback=self._get_human_feedback_enabled()
                )

                if plan.get("error"):
                    return json.dumps({
                        "status": "error",
                        "message": plan.get("error")
                    })

                # Save
                output_path = self.orch.base_dir / "plan.json"
                with open(output_path, 'w') as f:
                    json.dump(plan, f, indent=2)

                # Generate HTML
                from .html_generator import HTMLReportGenerator
                html_path = self.orch.base_dir / "plan.html"
                generator = HTMLReportGenerator(self.orch.planner.state)
                generator.generate(str(html_path))

                return json.dumps({
                    "status": "success",
                    "iteration": plan.get('iteration'),
                    "num_experiments": len(plan.get('proposed_experiments', [])),
                    "output_path": str(output_path),
                    "html_report": str(html_path),
                    "hint": "Use generate_implementation_code() or refine_implementation_code() to update executable code for the adjusted plan."
                })

            except Exception as e:
                logging.error(f"Plan adjustment error: {e}", exc_info=True)
                return json.dumps({
                    "status": "error",
                    "message": str(e)
                })

        self._register_tool(
            func=adjust_plan_for_constraints,
            name="adjust_plan_for_constraints",
            description=(
                "Adjusts the experimental plan when implementation or instrument "
                "constraints make the current plan impractical BEFORE running the experiment. "
                "Use when protocol generation reveals incompatibilities (e.g., pipette type "
                "vs plate layout, equipment limitations, reagent availability). "
                "Does NOT increment iteration or log as experimental results. "
                "Use refine_plan_with_results() instead when adjusting based on actual experimental outcomes."
            ),
            parameters={
                "constraint_description": {
                    "type": "string",
                    "description": (
                        "Description of the implementation constraint or instrument "
                        "incompatibility that requires plan adjustment. Include what "
                        "the constraint is, why it conflicts with the current plan, "
                        "and any proposed resolution if known."
                    )
                }
            },
            required=["constraint_description"]
        )

        # 5. REFINE IMPLEMENTATION CODE (based on refined plan)
        def refine_implementation_code():
            """
            Updates implementation code for the most recently refined plan.
            Use after refine_plan_with_results() to add/update executable code.
            """
            
            if not self.orch.planner.state or not self.orch.planner.state.get("current_plan"):
                return json.dumps({
                    "status": "error",
                    "message": "No active plan. Refine a plan first using refine_plan_with_results()"
                })
            
            current_plan = self.orch.planner.state["current_plan"]
            
            print(f"  ⚡ Tool: Refining implementation code for iteration {current_plan.get('iteration')}...")
            
            try:
                updated_plan = self.orch.planner.refine_implementation_code(
                    plan=current_plan,
                    enable_human_feedback=self._get_human_feedback_enabled()
                )
                
                if updated_plan.get("error"):
                    return json.dumps({
                        "status": "error",
                        "message": updated_plan.get("error")
                    })
                
                # Save
                output_path = self.orch.base_dir / "plan_refined.json"
                with open(output_path, 'w') as f:
                    json.dump(updated_plan, f, indent=2)
                
                # Regenerate HTML
                from .html_generator import HTMLReportGenerator
                html_path = self.orch.base_dir / "plan_refined.html"
                generator = HTMLReportGenerator(self.orch.planner.state)
                generator.generate(str(html_path))
                
                # Save scripts
                final_out = str(self.orch.base_dir / "output_scripts")
                print(f"\n--- Saving Scripts to: {final_out} ---")
                write_experiments_to_disk(updated_plan, final_out)
                
                return json.dumps({
                    "status": "success",
                    "message": "Implementation code updated",
                    "output_path": str(output_path),
                    "html_report": str(html_path),
                    "scripts_saved_to": final_out
                })
                
            except Exception as e:
                logging.error(f"Code refinement error: {e}", exc_info=True)
                return json.dumps({
                    "status": "error",
                    "message": str(e)
                })
        
        self._register_tool(
            func=refine_implementation_code,
            name="refine_implementation_code",
            description=(
                "Updates implementation code for the most recently refined plan. "
                "Maps refined experimental steps to executable code. "
                "Use after refine_plan_with_results() once the scientific strategy is approved."
            ),
            parameters={},
            required=[]
        )
        
        def analyze_file(
                file_path: str,
                extraction_goal: str = None,
                force_regenerate: bool = False,
                inputs: list[str] = None,
                targets: list[str] = None):
            """
            Analyzes a raw data file (CSV/XLSX) to extract metrics.
            
            Args:
                file_path: Path to data file
                extraction_goal: What to extract
                force_regenerate: If True, regenerates analysis script even if one exists.
                inputs: List of column names to treat as INPUT parameters for optimization
                targets: List of column names to treat as OPTIMIZATION TARGETS
            """
            print(f"  ⚡ Tool: Analyzing {file_path}...")

            resolved_path, error = self._resolve_data_path(file_path)
            if error:
                return error
            file_path = resolved_path

            # Resolve absolute path for tracking
            file_path_abs = str(Path(file_path).resolve())
            
            #  Build schema-aware extraction goal
            enhanced_objective = extraction_goal or ""
            # Always include the campaign objective so the scalarizer knows
            # what physically meaningful targets to derive
            if self.orch.objective and self.orch.objective != "Undefined Research Goal":
                enhanced_objective = (
                    f"Research objective: {self.orch.objective}\n\n{enhanced_objective}"
                ).strip()

            if inputs and targets:
                # User explicitly specified schema - incorporate into the objective query
                schema_instruction = f"""
        REQUIRED OUTPUT SCHEMA:
        - INPUT PARAMETERS (for optimization): {inputs}
        - TARGET METRICS (to optimize): {targets}

        Extract EXACTLY these columns from the data. Each row should contain values for all input parameters and all target metrics.
        For multi-objective optimization, we need BOTH targets: {targets}
        """
                enhanced_objective = f"{enhanced_objective}\n\n{schema_instruction}".strip()
                print(f"    📊 User-specified schema:")
                print(f"       Inputs: {inputs}")
                print(f"       Targets: {targets}")
            
            # Determine script to use.
            # Strategy: if a locked script exists, always try it first. If it
            # fails on the new data (e.g., different columns), auto-regenerate.
            # This prevents false schema-change triggers from LLM-supplied
            # target names that don't exactly match stored names.
            has_locked_script = (
                self.orch.active_scalarizer_script
                and Path(self.orch.active_scalarizer_script).exists()
            )
            if force_regenerate:
                script_to_use = None
                print(f"    🔄 Force regenerate: Creating new analysis script")
            elif has_locked_script:
                script_to_use = self.orch.active_scalarizer_script
                print(f"    (Consistency Mode: Using cached script)")
            else:
                script_to_use = None
                print(f"    (Discovery Mode: Generating new script)")

            # Pass schema to experiment context
            current_plan = self.orch.planner.state.get("current_plan", {})
            exp_context = current_plan.get("proposed_experiments", [{}])[0] if current_plan else {}

            # Inject schema requirements into context (only when generating new script)
            role_hints = None
            if inputs and targets and not has_locked_script:
                exp_context = exp_context.copy() if exp_context else {}
                exp_context["_schema_requirements"] = {
                    "input_columns": inputs,
                    "target_columns": targets,
                    "optimization_type": "multi-objective" if len(targets) > 1 else "single-objective"
                }
                role_hints = {"inputs": inputs, "targets": targets}

            try:
                res = self.orch.scalarizer.scalarize(
                    data_path=file_path,
                    objective_query=enhanced_objective,
                    reuse_script_path=script_to_use,
                    experiment_context=exp_context,
                    enable_human_review=self._get_human_feedback_enabled(),
                    column_role_hints=role_hints
                )

                if res["status"] != "success":
                    hint = "Try force_regenerate=True if the data format has changed"
                    if script_to_use:
                        hint = (
                            "The locked analysis script failed on this file. "
                            "If the data format has changed, use force_regenerate=True "
                            "to create a new script."
                        )
                    return json.dumps({
                        "status": "error",
                        "message": res.get('error', 'Analysis failed'),
                        "hint": hint
                    })

                # Validate sidecar conditions match script output
                # If the script hardcoded values from a different file's sidecar,
                # the output will have wrong condition values. Detect and auto-regenerate.
                if script_to_use and not force_regenerate:
                    sidecar_path = Path(file_path).with_suffix('.json')
                    if sidecar_path.exists():
                        try:
                            with open(sidecar_path) as _sc:
                                sidecar_data = json.load(_sc)
                            metrics_to_check = res["metrics"]
                            if isinstance(metrics_to_check, list):
                                metrics_to_check = metrics_to_check[0] if metrics_to_check else {}
                            mismatched = []
                            for key, expected_val in sidecar_data.items():
                                if key in metrics_to_check:
                                    actual_val = metrics_to_check[key]
                                    if isinstance(expected_val, (int, float)) and isinstance(actual_val, (int, float)):
                                        if abs(actual_val - expected_val) > 1e-6:
                                            mismatched.append(f"{key}: expected {expected_val}, got {actual_val}")
                            if mismatched:
                                print(f"    ⚠️  Sidecar mismatch detected (script has hardcoded values):")
                                for m in mismatched:
                                    print(f"       {m}")
                                print(f"    🔄 Auto-regenerating script...")
                                res = self.orch.scalarizer.scalarize(
                                    data_path=file_path,
                                    objective_query=enhanced_objective,
                                    reuse_script_path=None,
                                    experiment_context=exp_context,
                                    enable_human_review=self._get_human_feedback_enabled(),
                                    column_role_hints=role_hints
                                )
                                if res["status"] != "success":
                                    return json.dumps({
                                        "status": "error",
                                        "message": res.get('error', 'Regeneration failed'),
                                    })
                                force_regenerate = True  # ensure script lock updates below
                        except Exception:
                            pass  # If sidecar can't be read, skip validation

                if not self.orch.active_scalarizer_script or force_regenerate:
                    self.orch.active_scalarizer_script = res["source_script"]
                    print(f"    ✅ Analysis Logic Locked: {Path(self.orch.active_scalarizer_script).name}")

                # Merge sidecar conditions into scalarizer output.
                # analyze_batch merges conditions externally (line ~1584),
                # but analyze_file did not — so a batch-generated script
                # that only outputs target metrics (e.g. Peak_Absorbance)
                # would miss input parameters (e.g. temperature, pH) that
                # live in the sidecar JSON, causing a column mismatch on
                # CSV append.
                sidecar_merge_path = Path(file_path).with_suffix('.json')
                if sidecar_merge_path.exists():
                    try:
                        with open(sidecar_merge_path, 'r') as _sc:
                            sidecar_conds = json.load(_sc)
                        if isinstance(sidecar_conds, dict):
                            scalar_conds = {
                                k: v for k, v in sidecar_conds.items()
                                if isinstance(v, (int, float, str))
                            }
                            if scalar_conds:
                                raw = res["metrics"]
                                if isinstance(raw, list):
                                    res["metrics"] = [{**row, **scalar_conds} for row in raw]
                                elif isinstance(raw, dict):
                                    res["metrics"] = {**raw, **scalar_conds}
                                print(f"    📎 Merged sidecar conditions: {list(scalar_conds.keys())}")
                    except Exception as e:
                        logging.warning(f"Could not merge sidecar conditions: {e}")

                # Handle both single-row and multi-row results
                metrics = res["metrics"]
                
                if isinstance(metrics, list):
                    df_new = pd.DataFrame(metrics)
                    print(f"    📊 Processing {len(df_new)} data points from multi-well experiment")
                elif isinstance(metrics, dict):
                    df_new = pd.DataFrame([metrics])
                else:
                    return json.dumps({
                        "status": "error",
                        "message": f"Unexpected metrics format: {type(metrics)}"
                    })
                
                # DEDUPLICATION - Content-based tracking
                # Compute current file hash
                current_hash = self._compute_file_hash(file_path)
                current_row_count = len(df_new)

                # Get previous tracking for this file (handle both old and new format)
                prev_tracking = self.orch.analyzed_files.get(file_path_abs, {})
                if isinstance(prev_tracking, dict):
                    prev_hash = prev_tracking.get('hash')
                    prev_row_count = prev_tracking.get('row_count', 0)
                else:
                    # Legacy format: just row count as int
                    prev_hash = None
                    prev_row_count = prev_tracking

                # Check for duplicate content across different filenames
                for tracked_path, tracking_info in self.orch.analyzed_files.items():
                    if tracked_path == file_path_abs:
                        continue  # Skip self
                    tracked_hash = tracking_info.get('hash') if isinstance(tracking_info, dict) else None
                    if tracked_hash and tracked_hash == current_hash:
                        print(f"    ⚠️  Duplicate content detected - matches: {Path(tracked_path).name}")
                        df_final = pd.read_csv(self.orch.bo_data_path) if self.orch.bo_data_path.exists() else pd.DataFrame()
                        return json.dumps({
                            "status": "warning",
                            "message": f"This file's content was already analyzed from '{Path(tracked_path).name}'",
                            "data_points_collected": len(df_final),
                            "rows_added": 0,
                            "optimization_ready": len(df_final) >= 3,
                            "hint": "Data already in optimization set. No action needed unless this is different data with identical content."
                        })

                # Determine what to process based on hash and row count
                if prev_hash is None:
                    # FIRST TIME analyzing this file
                    print(f"    ✨ First time analyzing this file")
                    df_to_append = df_new
                    num_new = len(df_new)

                elif prev_hash != current_hash:
                    # FILE CONTENT CHANGED - reprocess entirely
                    print(f"    🔄 File content changed (hash mismatch) - reprocessing entirely")
                    
                    # Remove old data from optimization_data.csv if it exists
                    if self.orch.bo_data_path.exists() and prev_row_count > 0:
                        try:
                            df_existing = pd.read_csv(self.orch.bo_data_path)
                            # Remove the last prev_row_count rows (assumes they're from this file)
                            if len(df_existing) >= prev_row_count:
                                df_existing = df_existing.iloc[:-prev_row_count]
                                df_existing.to_csv(self.orch.bo_data_path, index=False)
                                print(f"    🗑️  Removed {prev_row_count} old rows from optimization data")
                        except Exception as e:
                            logging.warning(f"Could not clean old data: {e}")
                    
                    df_to_append = df_new
                    num_new = len(df_new)

                elif current_row_count > prev_row_count:
                    # ROWS APPENDED - process only new rows
                    df_new_only = df_new.iloc[prev_row_count:]
                    num_skipped = prev_row_count
                    num_new = len(df_new_only)
                    
                    if num_skipped > 0:
                        print(f"    🔍 Skipped {num_skipped} previously analyzed row(s)")
                    print(f"    ✅ Adding {num_new} NEW row(s)")
                    
                    df_to_append = df_new_only

                elif current_row_count == prev_row_count:
                    # prev_hash == current_hash (guaranteed by earlier elif)
                    schema_changed = (
                        inputs and targets and (
                            set(inputs) != set(self.orch.expected_input_columns or [])
                            or set(targets) != set(self.orch.expected_target_columns or [])
                        )
                    )
                    if schema_changed:
                        # Same data, new schema — reprocess with updated columns
                        print(f"    🔄 Schema changed — reprocessing with new inputs/targets")
                        df_to_append = df_new
                        num_new = len(df_new)
                        # Clear existing optimization data (schema mismatch)
                        if self.orch.bo_data_path.exists():
                            backup = self.orch.bo_data_path.with_suffix('.csv.backup')
                            self.orch.bo_data_path.rename(backup)
                            print(f"    ⚠️  Old optimization data backed up (schema change)")
                    else:
                        # TRULY UNCHANGED
                        print(f"    ℹ️  File unchanged (same content hash)")
                        df_final = pd.read_csv(self.orch.bo_data_path) if self.orch.bo_data_path.exists() else pd.DataFrame()

                        return json.dumps({
                            "status": "success",
                            "message": "File already analyzed - no changes detected",
                            "data_points_collected": len(df_final),
                            "rows_added": 0,
                            "optimization_ready": len(df_final) >= 3
                        })

                else:
                    # FEWER ROWS - file was truncated/replaced
                    print(f"    ⚠️  File has fewer rows ({current_row_count} < {prev_row_count}) - reprocessing")
                    
                    # Remove old data
                    if self.orch.bo_data_path.exists() and prev_row_count > 0:
                        try:
                            df_existing = pd.read_csv(self.orch.bo_data_path)
                            if len(df_existing) >= prev_row_count:
                                df_existing = df_existing.iloc[:-prev_row_count]
                                df_existing.to_csv(self.orch.bo_data_path, index=False)
                                print(f"    🗑️  Removed {prev_row_count} old rows from optimization data")
                        except Exception as e:
                            logging.warning(f"Could not clean old data: {e}")
                    
                    df_to_append = df_new
                    num_new = len(df_new)

                # Schema enforcement BEFORE saving
                all_cols = list(df_to_append.columns)

                # Case 1: Agent explicitly provided schema (Enables MOO)
                if inputs and targets:
                    # Validate that requested columns exist in the extracted data
                    missing_inputs = [c for c in inputs if c not in all_cols]
                    missing_targets = [t for t in targets if t not in all_cols]
                    
                    if missing_inputs or missing_targets:
                        # Try fuzzy matching for column names
                        available_cols = all_cols
                        suggestions = {}
                        
                        for missing in missing_inputs + missing_targets:
                            # Simple fuzzy match: find columns containing similar substrings
                            matches = [c for c in available_cols if missing.lower().replace('_', '') in c.lower().replace('_', '') 
                                    or c.lower().replace('_', '') in missing.lower().replace('_', '')]
                            if matches:
                                suggestions[missing] = matches
                        
                        return json.dumps({
                            "status": "schema_mismatch",
                            "message": (
                                f"The analysis script could not produce the requested columns. "
                                f"Missing inputs: {missing_inputs or 'none'}. "
                                f"Missing targets: {missing_targets or 'none'}. "
                                f"Available columns from extraction: {all_cols}."
                            ),
                            "missing_inputs": missing_inputs,
                            "missing_targets": missing_targets,
                            "available_columns": all_cols,
                            "suggestions": suggestions if suggestions else None,
                            "recovery_options": [
                                "Retry with corrected column names from available_columns",
                                "Use force_regenerate=True with an updated extraction_goal",
                                "Choose different inputs/targets from the available columns"
                            ]
                        })
                    
                    self.orch.expected_input_columns = inputs
                    self.orch.expected_target_columns = targets
                    # Capture optimization direction and input types from scalarizer
                    column_roles = res.get("column_roles", {})
                    opt_dir = column_roles.get("optimization_direction", {})
                    if opt_dir:
                        self.orch.target_directions = opt_dir
                    self._capture_input_types(column_roles, inputs)
                    print(f"    📊 Schema Enforced (User-Specified):")
                    print(f"       Inputs: {self.orch.expected_input_columns}")
                    print(f"       Targets: {self.orch.expected_target_columns}")
                    if self.orch.target_directions:
                        print(f"       Directions: {self.orch.target_directions}")

                # Case 2: Schema already established from previous analysis
                elif self.orch.expected_input_columns and self.orch.expected_target_columns:
                    # Still capture direction if scalarizer provided it and we don't have one yet
                    column_roles = res.get("column_roles", {})
                    if not self.orch.target_directions:
                        opt_dir = column_roles.get("optimization_direction", {})
                        if opt_dir:
                            self.orch.target_directions = opt_dir
                    if not getattr(self.orch, "expected_input_types", None):
                        self._capture_input_types(column_roles, self.orch.expected_input_columns)
                    print(f"    📊 Schema Enforced (From Previous Analysis):")
                    print(f"       Inputs: {self.orch.expected_input_columns}")
                    print(f"       Targets: {self.orch.expected_target_columns}")
                
                # Case 3: No user schema — use scalarizer's column_roles classification
                else:
                    column_roles = res.get("column_roles", {})
                    proposed_inputs = column_roles.get("inputs", [])
                    proposed_targets = column_roles.get("targets", [])

                    if proposed_inputs and proposed_targets:
                        # Validate proposed columns exist in extracted data
                        missing = [c for c in proposed_inputs + proposed_targets if c not in all_cols]
                        if missing:
                            print(f"    ⚠️  Scalarizer classification references missing columns: {missing}")
                            # Fall through to schema_required below
                            proposed_inputs, proposed_targets = [], []
                        else:
                            reasoning = column_roles.get("reasoning", "")
                            opt_dir = column_roles.get("optimization_direction", {})
                            print(f"    🔬 Scalarizer classified columns:")
                            print(f"       Inputs: {proposed_inputs}")
                            print(f"       Targets: {proposed_targets}")
                            if opt_dir:
                                print(f"       Directions: {opt_dir}")
                            if reasoning:
                                print(f"       Reasoning: {reasoning}")

                            if self.orch.autonomy_level == "CO_PILOT":
                                # Return proposal for user confirmation
                                n_data = len(df_to_append)
                                return json.dumps({
                                    "status": "schema_proposed",
                                    "inputs": proposed_inputs,
                                    "targets": proposed_targets,
                                    "optimization_direction": opt_dir,
                                    "reasoning": reasoning,
                                    "data_points": n_data,
                                    "message": "Scalarizer proposes this classification. Confirm or adjust.",
                                    "available_columns": all_cols
                                })
                            else:
                                # SUPERVISED/AUTONOMOUS: accept directly
                                self.orch.expected_input_columns = proposed_inputs
                                self.orch.expected_target_columns = proposed_targets
                                if opt_dir:
                                    self.orch.target_directions = opt_dir
                                self._capture_input_types(column_roles, proposed_inputs)
                                print(f"    ✅ Schema auto-accepted: inputs={proposed_inputs}, targets={proposed_targets}")

                    # Targets found but no inputs — measurement-only data (e.g., spectra)
                    if proposed_targets and not proposed_inputs:
                        n_data = len(df_to_append)
                        return json.dumps({
                            "status": "inputs_required",
                            "message": (
                                "The data file contains measurement data but no experimental conditions. "
                                "Input parameters (e.g., temperature, pH, concentration) are needed for optimization."
                            ),
                            "targets_found": proposed_targets,
                            "reasoning": column_roles.get("reasoning", ""),
                            "data_points": n_data,
                            "options": [
                                "Provide a metadata JSON sidecar file with experimental conditions",
                                "Manually specify input parameter values for this data file",
                                "Re-call analyze_file with inputs=[...] listing the parameter names to add"
                            ]
                        })

                    # Fallback: scalarizer didn't classify or classification was invalid
                    if not proposed_inputs or not proposed_targets:
                        print(f"    ⚠️  No schema specified. Extracted columns: {all_cols}")
                        data_preview = df_to_append.head(3).to_dict(orient='records')
                        col_stats = {}
                        for col in all_cols:
                            if pd.api.types.is_numeric_dtype(df_to_append[col]):
                                col_stats[col] = {
                                    "type": "numeric",
                                    "unique": int(df_to_append[col].nunique()),
                                    "min": float(df_to_append[col].min()),
                                    "max": float(df_to_append[col].max()),
                                }
                            else:
                                col_stats[col] = {"type": "non-numeric", "unique": int(df_to_append[col].nunique())}
                        n_data = len(df_to_append)
                        numeric_cols = [c for c in all_cols if pd.api.types.is_numeric_dtype(df_to_append[c])]
                        return json.dumps({
                            "status": "schema_required",
                            "message": "Could not auto-classify columns. Re-call analyze_file with explicit inputs and targets.",
                            "available_columns": all_cols,
                            "column_stats": col_stats,
                            "data_preview": data_preview,
                            "data_points": n_data,
                            "objective": self.orch.objective or "Not set",
                            "hint": (
                                "Use the objective, column names, and data preview to decide: "
                                "which columns are controllable INPUT parameters (experimentally set) "
                                "and which are measured TARGET metrics (outcomes to optimize). "
                                "Non-numeric columns (e.g., Sample_ID, Notes) should be excluded. "
                                "Then call: analyze_file(file_path=..., inputs=[...], targets=[...])"
                            ),
                            "objective_count_guidance": self._build_objective_guidance(n_data, numeric_cols)
                        })
                
                # FILTER TO MATCH CSV SCHEMA
                # The scalarizer may output extra metrics beyond what the
                # optimization CSV tracks.  Use the existing CSV columns as
                # ground truth (they may be wider than expected_target_columns
                # if targets were narrowed via run_optimization).  For new
                # CSVs, fall back to the expected schema.
                if self.orch.bo_data_path.exists():
                    df_existing = pd.read_csv(self.orch.bo_data_path)
                    ref_cols = list(df_existing.columns)
                else:
                    df_existing = None
                    if self.orch.expected_input_columns and self.orch.expected_target_columns:
                        ref_cols = self.orch.expected_input_columns + self.orch.expected_target_columns
                    else:
                        ref_cols = None

                if ref_cols:
                    available = [c for c in ref_cols if c in df_to_append.columns]
                    extra = [c for c in df_to_append.columns if c not in ref_cols]
                    missing = [c for c in ref_cols if c not in df_to_append.columns]
                    if extra:
                        print(f"    📎 Dropping extra columns: {extra}")
                    if missing:
                        print(f"    ⚠️  Missing expected columns: {missing}")
                    if available:
                        df_to_append = df_to_append[available]

                # SCHEMA ENFORCEMENT ON SAVE
                if df_existing is not None:
                    if set(df_to_append.columns) != set(df_existing.columns):
                        return json.dumps({
                            "status": "error",
                            "message": "Schema mismatch detected",
                            "expected_columns": list(df_existing.columns),
                            "received_columns": list(df_to_append.columns),
                            "hint": "All data must have same structure. Use reset_analysis_logic to start fresh."
                        })

                    df_to_append = df_to_append[df_existing.columns]
                    df_to_append.to_csv(self.orch.bo_data_path, mode='a', header=False, index=False)
                else:
                    df_to_append.to_csv(self.orch.bo_data_path, mode='w', header=True, index=False)
                
                # Update tracking
                self.orch.analyzed_files[file_path_abs] = {
                    'row_count': current_row_count,
                    'hash': current_hash,
                    'timestamp': datetime.now().isoformat()
                }
                with open(self.orch.analyzed_files_path, 'w') as f:
                    json.dump(self.orch.analyzed_files, f, indent=2)
                
                df_final = pd.read_csv(self.orch.bo_data_path)
                data_count = len(df_final)
                
                return json.dumps({
                    "status": "success",
                    "data_points_collected": data_count,
                    "rows_added": num_new,
                    "optimization_ready": data_count >= 3,
                    "inputs": self.orch.expected_input_columns,
                    "targets": self.orch.expected_target_columns
                })
                
            except Exception as e:
                logging.error(f"Analyze file error: {e}", exc_info=True)
                return json.dumps({
                    "status": "error",
                    "message": str(e)
                })
        
        self._register_tool(
            func=analyze_file,
            name="analyze_file",
            description=(
                "Analyzes raw data files (CSV/XLSX/TXT) to extract scalar metrics. "
                "Automatically generates analysis code on first use, then reuses it for consistency. "
                "Results are appended to optimization dataset."
            ),
            parameters={
                "file_path": {
                    "type": "string",
                    "description": "Path to the data file to analyze (e.g., 'results/run_001.csv')"
                },
                "extraction_goal": {
                    "type": "string",
                    "description": "Natural language description of what to extract (e.g., 'Calculate peak area and retention time')"
                },
                "force_regenerate": {
                    "type": "boolean",
                    "description": (
                        "If true, generates new analysis script even if one exists. "
                        "Use when analysis requirements change (e.g., switching from single-row to multi-row extraction, "
                        "or changing which metrics to extract). Default: false"
                    )
                },
                "inputs": {
                    "type": "array", 
                    "items": {"type": "string"},
                    "description": "List of column names to treat as INPUT parameters"
                },
                "targets": {
                    "type": "array", 
                    "items": {"type": "string"}, 
                    "description": "List of column names to treat as OPTIMIZATION TARGETS"
                }
            },
            required=["file_path"]
        )
        
        # 6b. ANALYZE BATCH
        def analyze_batch(
                file_paths: list[str],
                extraction_goal: str = None,
                conditions: str = None,
                conditions_file: str = None,
                inputs: list[str] = None,
                targets: list[str] = None,
                force_regenerate: bool = False):
            """
            Analyzes multiple raw data files in a single call.
            Runs the scalarizer on each file, merges with experimental conditions,
            and appends all results to the optimization dataset.
            """
            print(f"  ⚡ Tool: Analyzing batch of {len(file_paths)} files...")

            # --- 1. Resolve all file paths first ---
            resolved_paths = []
            resolve_errors = []
            for fp in file_paths:
                resolved, err = self._resolve_data_path(fp)
                if err:
                    resolve_errors.append({"file": fp, "error": "File not found"})
                else:
                    resolved_paths.append(resolved)

            if not resolved_paths:
                return json.dumps({
                    "status": "error",
                    "message": "None of the provided file paths could be resolved.",
                    "errors": resolve_errors
                })

            # --- 2. Parse conditions ---
            file_conditions = {}  # resolved_path -> {param: value}

            raw_conditions = None
            if conditions:
                try:
                    raw_conditions = json.loads(conditions)
                except json.JSONDecodeError as e:
                    return json.dumps({
                        "status": "error",
                        "message": f"Could not parse conditions JSON: {e}",
                        "expected_formats": [
                            '{"filename.csv": {"temperature": 300, "pH": 7.0}, ...}',
                            '[{"temperature": 300}, {"temperature": 350}]'
                        ]
                    })
            elif conditions_file:
                cf_resolved, cf_err = self._resolve_data_path(conditions_file)
                if cf_err:
                    return json.dumps({
                        "status": "error",
                        "message": f"Conditions file not found: {conditions_file}"
                    })
                try:
                    with open(cf_resolved, 'r') as f:
                        raw_conditions = json.load(f)
                except Exception as e:
                    return json.dumps({
                        "status": "error",
                        "message": f"Could not read conditions file: {e}"
                    })

            if raw_conditions is not None:
                if isinstance(raw_conditions, list):
                    # Positional matching
                    if len(raw_conditions) != len(resolved_paths):
                        return json.dumps({
                            "status": "error",
                            "message": (
                                f"Conditions list has {len(raw_conditions)} entries "
                                f"but {len(resolved_paths)} files were resolved. Counts must match."
                            )
                        })
                    for rpath, cond in zip(resolved_paths, raw_conditions):
                        if isinstance(cond, dict):
                            file_conditions[rpath] = cond
                elif isinstance(raw_conditions, dict):
                    # Key matching by filename or stem
                    unmatched_keys = []
                    for key, cond in raw_conditions.items():
                        matched = False
                        for rpath in resolved_paths:
                            fname = Path(rpath).name
                            stem = Path(rpath).stem
                            if key in (fname, stem):
                                file_conditions[rpath] = cond
                                matched = True
                                break
                        if not matched:
                            unmatched_keys.append(key)
                    if unmatched_keys and not file_conditions:
                        return json.dumps({
                            "status": "error",
                            "message": (
                                f"None of the condition keys matched any file. "
                                f"Unmatched keys: {unmatched_keys}. "
                                f"Resolved filenames: {[Path(p).name for p in resolved_paths]}"
                            )
                        })
                    if unmatched_keys:
                        print(f"    ⚠️  Unmatched condition keys (ignored): {unmatched_keys}")

            # --- 3. Process each file through scalarizer ---
            enhanced_objective = extraction_goal or ""
            if self.orch.objective and self.orch.objective != "Undefined Research Goal":
                enhanced_objective = (
                    f"Research objective: {self.orch.objective}\n\n{enhanced_objective}"
                ).strip()
            if inputs and targets:
                schema_instruction = (
                    f"\nREQUIRED OUTPUT SCHEMA:\n"
                    f"- INPUT PARAMETERS: {inputs}\n"
                    f"- TARGET METRICS: {targets}\n"
                    f"Extract EXACTLY these columns from the data."
                )
                enhanced_objective = f"{enhanced_objective}\n{schema_instruction}".strip()

            current_plan = self.orch.planner.state.get("current_plan", {})
            exp_context = current_plan.get("proposed_experiments", [{}])[0] if current_plan else {}

            if inputs and targets:
                exp_context = exp_context.copy() if exp_context else {}
                exp_context["_schema_requirements"] = {
                    "input_columns": inputs,
                    "target_columns": targets,
                    "optimization_type": "multi-objective" if len(targets) > 1 else "single-objective"
                }

            role_hints = {"inputs": inputs, "targets": targets} if inputs and targets else None

            script_to_use = None
            if not force_regenerate:
                script_to_use = self.orch.active_scalarizer_script if (
                    self.orch.active_scalarizer_script and Path(self.orch.active_scalarizer_script).exists()
                ) else None

            results = []
            errors = list(resolve_errors)
            first_column_roles = None

            for rpath in resolved_paths:
                fname = Path(rpath).name
                try:
                    res = self.orch.scalarizer.scalarize(
                        data_path=rpath,
                        objective_query=enhanced_objective,
                        reuse_script_path=script_to_use,
                        experiment_context=exp_context,
                        enable_human_review=False,
                        column_role_hints=role_hints
                    )

                    if res.get("status") != "success":
                        errors.append({"file": fname, "error": res.get("error", "Scalarizer failed")})
                        continue

                    # Lock script after first success
                    if not script_to_use and res.get("source_script"):
                        script_to_use = res["source_script"]
                        self.orch.active_scalarizer_script = script_to_use
                        print(f"    🔒 Script locked: {Path(script_to_use).name}")

                    if first_column_roles is None:
                        first_column_roles = res.get("column_roles", {})

                    # Extract metrics row
                    metrics = res.get("metrics", {})
                    if isinstance(metrics, list):
                        if len(metrics) == 1:
                            row = metrics[0]
                        else:
                            # Multi-row from a single spectrum file — take all rows
                            # but this is unusual; log a warning
                            print(f"    ⚠️  {fname}: scalarizer returned {len(metrics)} rows (expected 1)")
                            row = metrics[0]
                    elif isinstance(metrics, dict):
                        row = metrics
                    else:
                        errors.append({"file": fname, "error": f"Unexpected metrics type: {type(metrics)}"})
                        continue

                    # Merge conditions with scalarizer output.
                    # External conditions take priority — they overwrite any
                    # values the script may have hardcoded from the first sidecar.
                    cond = file_conditions.get(rpath, {})
                    merged = {**row, **cond}
                    results.append({"file": fname, "row": merged, "path": rpath})

                    print(f"    ✅ {fname}: {len(merged)} columns extracted")

                except Exception as e:
                    logging.error(f"Error processing {fname}: {e}", exc_info=True)
                    errors.append({"file": fname, "error": str(e)})

            if not results:
                return json.dumps({
                    "status": "error",
                    "message": "All files failed during processing.",
                    "errors": errors
                })

            # --- 4. Determine schema ---
            all_keys = list(results[0]["row"].keys())
            condition_keys = set()
            if file_conditions:
                for cond in file_conditions.values():
                    condition_keys.update(cond.keys())
                condition_keys = sorted(condition_keys)

            if inputs and targets:
                # User-specified schema
                self.orch.expected_input_columns = inputs
                self.orch.expected_target_columns = targets
                self._capture_input_types(first_column_roles or {}, inputs)
            elif condition_keys:
                # Derive: inputs = condition keys, targets = remaining keys
                proposed_targets = first_column_roles.get("targets", []) if first_column_roles else []
                scalarizer_keys = [k for k in all_keys if k not in condition_keys]
                if not proposed_targets:
                    proposed_targets = scalarizer_keys
                self.orch.expected_input_columns = list(condition_keys)
                self.orch.expected_target_columns = proposed_targets
                self._capture_input_types(first_column_roles or {}, list(condition_keys))
            else:
                # No conditions at all — check if scalarizer found inputs
                proposed_inputs = first_column_roles.get("inputs", []) if first_column_roles else []
                proposed_targets = first_column_roles.get("targets", []) if first_column_roles else []
                if proposed_inputs and proposed_targets:
                    self.orch.expected_input_columns = proposed_inputs
                    self.orch.expected_target_columns = proposed_targets
                    self._capture_input_types(first_column_roles or {}, proposed_inputs)
                else:
                    # Return inputs_required with example template
                    file_names = [r["file"] for r in results]
                    targets_found = proposed_targets or all_keys
                    example = {fn: {"parameter_1": "value", "parameter_2": "value"} for fn in file_names}
                    return json.dumps({
                        "status": "inputs_required",
                        "message": (
                            f"Extracted targets from {len(results)} files but no experimental "
                            f"conditions were found. Input parameters are needed for optimization."
                        ),
                        "files": file_names,
                        "targets_found": targets_found,
                        "example_conditions": example,
                        "options": [
                            "Re-call analyze_batch with conditions='{...}' mapping filenames to parameter values",
                            "Re-call analyze_batch with conditions_file='path/to/conditions.json'",
                            "Place sidecar JSON files next to each data file (e.g., spectrum_300C.json)"
                        ]
                    })

            print(f"    📊 Batch schema: inputs={self.orch.expected_input_columns}, targets={self.orch.expected_target_columns}")

            # --- 5. Validate condition key consistency across files ---
            files_missing_conditions = []
            if condition_keys:
                for r in results[:]:  # iterate over copy
                    missing_cond_keys = [k for k in condition_keys if k not in r["row"] or r["row"][k] is None]
                    if missing_cond_keys:
                        files_missing_conditions.append(r["file"])
                        results.remove(r)

            if not results and files_missing_conditions:
                example = {fn: {k: "value" for k in condition_keys} for fn in files_missing_conditions}
                return json.dumps({
                    "status": "inputs_required",
                    "message": "All files are missing experimental conditions.",
                    "files_missing_conditions": files_missing_conditions,
                    "example_conditions": example,
                    "options": [
                        "Re-call analyze_batch with conditions='{...}' providing values for all files",
                        "Re-call analyze_batch with conditions_file='path/to/conditions.json'"
                    ]
                })

            # --- 6. Build DataFrame and append to optimization_data.csv ---
            expected_cols = self.orch.expected_input_columns + self.orch.expected_target_columns

            # Filter rows to expected columns only
            clean_rows = []
            for r in results:
                row_data = {}
                for col in expected_cols:
                    row_data[col] = r["row"].get(col)
                clean_rows.append(row_data)

            df_batch = pd.DataFrame(clean_rows)

            if self.orch.bo_data_path.exists():
                df_existing = pd.read_csv(self.orch.bo_data_path)
                if set(df_batch.columns) != set(df_existing.columns):
                    if inputs or targets:
                        # Schema intentionally changed by user — replace old data
                        print(f"    🔄 Schema changed (user-specified targets). Replacing old optimization data.")
                        backup = self.orch.bo_data_path.with_suffix('.csv.backup')
                        self.orch.bo_data_path.rename(backup)
                        # Clear file tracking so all files are re-counted
                        self.orch.analyzed_files = {}
                        df_batch.to_csv(self.orch.bo_data_path, mode='w', header=True, index=False)

                        for r in results:
                            file_path_abs = str(Path(r["path"]).resolve())
                            current_hash = self._compute_file_hash(r["path"])
                            self.orch.analyzed_files[file_path_abs] = {
                                'row_count': 1, 'hash': current_hash,
                                'timestamp': datetime.now().isoformat()
                            }
                        with open(self.orch.analyzed_files_path, 'w') as f:
                            json.dump(self.orch.analyzed_files, f, indent=2)

                        df_final = pd.read_csv(self.orch.bo_data_path)
                        return json.dumps({
                            "status": "success",
                            "files_processed": len(results),
                            "rows_added": len(results),
                            "data_points_collected": len(df_final),
                            "optimization_ready": len(df_final) >= 3,
                            "inputs": self.orch.expected_input_columns,
                            "targets": self.orch.expected_target_columns,
                            "note": "Schema changed — old data replaced with new target selection"
                        })
                    return json.dumps({
                        "status": "error",
                        "message": "Schema mismatch with existing optimization data.",
                        "expected_columns": list(df_existing.columns),
                        "received_columns": list(df_batch.columns),
                        "hint": "Use reset_analysis_logic to start fresh."
                    })
                df_batch = df_batch[df_existing.columns]
                df_batch.to_csv(self.orch.bo_data_path, mode='a', header=False, index=False)
            else:
                df_batch.to_csv(self.orch.bo_data_path, mode='w', header=True, index=False)

            # Track each file
            for r in results:
                file_path_abs = str(Path(r["path"]).resolve())
                current_hash = self._compute_file_hash(r["path"])
                self.orch.analyzed_files[file_path_abs] = {
                    'row_count': 1,
                    'hash': current_hash,
                    'timestamp': datetime.now().isoformat()
                }
            with open(self.orch.analyzed_files_path, 'w') as f:
                json.dump(self.orch.analyzed_files, f, indent=2)

            df_final = pd.read_csv(self.orch.bo_data_path)
            data_count = len(df_final)
            num_added = len(results)

            response = {
                "status": "success" if not errors and not files_missing_conditions else "partial_success",
                "files_processed": len(results),
                "rows_added": num_added,
                "data_points_collected": data_count,
                "optimization_ready": data_count >= 3,
                "inputs": self.orch.expected_input_columns,
                "targets": self.orch.expected_target_columns,
            }
            if errors:
                response["errors"] = errors
            if files_missing_conditions:
                example = {fn: {k: "value" for k in condition_keys} for fn in files_missing_conditions}
                response["files_missing_conditions"] = files_missing_conditions
                response["example_conditions"] = example

            return json.dumps(response)

        self._register_tool(
            func=analyze_batch,
            name="analyze_batch",
            description=(
                "Analyzes multiple data files (e.g., spectra, time series) in a single call. "
                "Runs the scalarizer on each file to extract target metrics, then merges with "
                "experimental conditions. All results are appended to the optimization dataset. "
                "Use instead of calling analyze_file repeatedly when files share the same structure."
            ),
            parameters={
                "file_paths": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": (
                        "List of data file paths to analyze "
                        "(e.g., ['data/spectrum_300C.csv', 'data/spectrum_350C.csv'])"
                    )
                },
                "extraction_goal": {
                    "type": "string",
                    "description": "What to extract from each file (e.g., 'Calculate peak area and FWHM')"
                },
                "conditions": {
                    "type": "string",
                    "description": (
                        "JSON string mapping filenames to experimental conditions. "
                        "Dict format: {\"spectrum_300C.csv\": {\"temperature\": 300}, ...}. "
                        "List format: [{\"temperature\": 300}, {\"temperature\": 350}] (positional). "
                        "Omit if sidecar JSONs exist next to each data file."
                    )
                },
                "conditions_file": {
                    "type": "string",
                    "description": (
                        "Path to a JSON file containing the conditions mapping "
                        "(same format as the conditions parameter)"
                    )
                },
                "inputs": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "List of column names to treat as INPUT parameters (overrides auto-detection)"
                },
                "targets": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "List of column names to treat as OPTIMIZATION TARGETS (overrides auto-detection)"
                },
                "force_regenerate": {
                    "type": "boolean",
                    "description": "If true, regenerates analysis script even if one exists. Default: false"
                }
            },
            required=["file_paths"]
        )

        # 7. RESET ANALYSIS LOGIC
        def reset_analysis_logic():
            """Resets the analysis script, optimization data, AND file tracking."""
            self.orch.active_scalarizer_script = None
            self.orch.expected_input_columns = None
            self.orch.expected_target_columns = []
            
            # Clear file tracking completely
            self.orch.analyzed_files = {}
            if self.orch.analyzed_files_path.exists():
                try:
                    self.orch.analyzed_files_path.unlink()
                    print(f"    🗑️  Cleared file tracking history")
                except Exception as e:
                    logging.warning(f"Could not delete analyzed_files.json: {e}")
            
            if self.orch.bo_data_path.exists():
                backup_path = self.orch.bo_data_path.with_suffix('.csv.backup')
                self.orch.bo_data_path.rename(backup_path)
                print(f"    ⚠️  Old data backed up to: {backup_path.name}")

            bo_history = self.orch.bo.history_file
            if bo_history.exists():
                backup = bo_history.with_suffix('.json.backup')
                bo_history.rename(backup)
                print(f"    ⚠️  BO history backed up to: {backup.name}")

            return json.dumps({
                "status": "success",
                "message": "Analysis logic reset. All files will be reprocessed fresh on next analyze_file call.",
                "hint": "Previous optimization data was backed up"
            })
        
        self._register_tool(
            func=reset_analysis_logic,
            name="reset_analysis_logic",
            description=(
                "Resets the locked analysis script and clears optimization data. "
                "Use this when the current analysis approach is fundamentally wrong. "
                "Previous data is backed up before deletion."
            ),
            parameters={},
            required=[]
        )
        
        # 8. RUN OPTIMIZATION
        def run_optimization(
            parallel_capable: bool = False,
            batch_size: int = None,
            physical_constraints: str = None,
            experimental_budget: int = None,
            targets: list[str] = None,
            strategy_hint: str = None
        ):
            """
            Runs Bayesian Optimization to suggest next parameters.
            Supports optional physical constraints for realizable batch design
            and optional experimental budget for exploration/exploitation control.
            """
            print(f"  ⚡ Tool: Running Bayesian Optimization...")
            
            # --- PRE-FLIGHT CHECKS --- 
            if not self.orch.active_scalarizer_script:
                return json.dumps({
                    "status": "error",
                    "message": "No analysis script locked yet",
                    "hint": "Run analyze_file on at least 3 data files first",
                    "workflow": "analyze_file (×3) → run_optimization"
                })
            
            if not self.orch.bo_data_path.exists():
                return json.dumps({
                    "status": "error",
                    "message": "No optimization_data.csv found",
                    "hint": "Run analyze_file to collect data points first"
                })
            
            try:
                df = pd.read_csv(self.orch.bo_data_path)
            except Exception as e:
                return json.dumps({
                    "status": "error",
                    "message": f"Failed to read optimization data: {e}",
                    "hint": "CSV may be corrupted. Check optimization_data.csv"
                })
            
            if len(df) < 3:
                return json.dumps({
                    "status": "error", 
                    "message": f"Insufficient data points: {len(df)}/3",
                    "hint": "Collect at least 3 experimental results before optimizing",
                    "current_data_count": len(df)
                })
            
            if not self.orch.expected_target_columns or not self.orch.expected_input_columns:
                return json.dumps({
                    "status": "error",
                    "message": "Schema not established",
                    "hint": "This shouldn't happen. Try reset_analysis_logic."
                })

            # TARGET NARROWING — allows switching from MOO to SOO
            # without re-running analyze_batch on every file.
            if targets:
                missing = [t for t in targets if t not in df.columns]
                if missing:
                    return json.dumps({
                        "status": "error",
                        "message": f"Requested targets not in data: {missing}",
                        "available_columns": list(df.columns)
                    })
                old_targets = self.orch.expected_target_columns
                self.orch.expected_target_columns = targets
                print(f"    🎯 Targets narrowed: {old_targets} → {targets}")

            # SCHEMA VALIDATION
            missing_targets = [t for t in self.orch.expected_target_columns if t not in df.columns]
            if missing_targets:
                return json.dumps({
                    "status": "error",
                    "message": f"Target columns missing from data: {missing_targets}",
                    "available_columns": list(df.columns)
                })
            
            missing_inputs = [c for c in self.orch.expected_input_columns if c not in df.columns]
            if missing_inputs:
                return json.dumps({
                    "status": "error",
                    "message": f"Input columns missing: {missing_inputs}",
                    "available_columns": list(df.columns)
                })
            
            critical_cols = self.orch.expected_input_columns + self.orch.expected_target_columns
            
            if df[critical_cols].isnull().any().any():
                return json.dumps({
                    "status": "error",
                    "message": "Missing values detected in optimization data",
                    "hint": "Ensure all data files were analyzed successfully",
                    "affected_rows": df[df[critical_cols].isnull().any(axis=1)].index.tolist()
                })
            
            # ============================================
            # BOUNDS & CONSTRAINTS CALCULATION
            # ============================================
            # Pull continuous-parameter bounds and categorical-parameter levels
            # from the planner's current_plan. The planner schema supports both
            # parameter_type="continuous" (with min_value/max_value) and
            # parameter_type="categorical" (with levels). Missing parameter_type
            # is treated as continuous for backward compatibility.
            scientific_bounds: Dict[str, tuple] = {}
            planner_levels: Dict[str, List[str]] = {}
            current_plan = self.orch.planner.state.get("current_plan", {})

            if current_plan and "proposed_experiments" in current_plan:
                for exp in current_plan["proposed_experiments"]:
                    for param in exp.get("optimization_params", []):
                        name = param.get("parameter_name")
                        if not name:
                            continue
                        ptype = param.get("parameter_type", "continuous")
                        if ptype == "categorical":
                            levels = param.get("levels") or []
                            if levels:
                                planner_levels[name] = [str(lv) for lv in levels]
                                print(f"  🔬 Scientific Constraint Found: {name} ∈ {planner_levels[name]}")
                            continue
                        min_v = param.get("min_value")
                        max_v = param.get("max_value")
                        if min_v is not None and max_v is not None:
                            scientific_bounds[name] = (float(min_v), float(max_v))
                            print(f"  🔬 Scientific Constraint Found: {name} must be between {min_v} and {max_v}")

            # Resolve input types: scalarizer is the source of truth on type;
            # planner is the source of truth on the level universe (so BO can
            # recommend levels not yet observed in the data).
            input_types_state = getattr(self.orch, "expected_input_types", None) or {}

            optimization_inputs: List[str] = []
            level_maps: Dict[str, List[str]] = {}
            type_conflict_warnings: List[str] = []

            for col in self.orch.expected_input_columns:
                if col not in df.columns:
                    print(f"  ⚠️ Skipping missing input column: {col}")
                    continue

                declared_type = input_types_state.get(col)
                planner_says_cat = col in planner_levels

                # Type resolution: scalarizer wins on type because it sees the
                # actual data file. The planner's declaration of `levels` is
                # only honored when the scalarizer agrees the column is
                # categorical (or hasn't classified it). This avoids silently
                # corrupting a continuous knob whose values happen to look
                # discrete in the observed data.
                if declared_type == "continuous":
                    if planner_says_cat:
                        type_conflict_warnings.append(
                            f"{col}: scalarizer classified as continuous, planner declared "
                            "categorical — honoring scalarizer (data shape wins). To force "
                            "categorical, fix the scalarizer's input_types classification."
                        )
                    is_categorical = False
                elif declared_type == "categorical" or planner_says_cat:
                    is_categorical = True
                else:
                    is_categorical = False

                if is_categorical:
                    observed = sorted(df[col].dropna().astype(str).unique().tolist())
                    if planner_levels.get(col):
                        # Planner is authoritative on the level universe.
                        # Observed values that aren't in the planner-declared
                        # set are a real misalignment (data needs re-encoding,
                        # or the planner's levels are wrong) — fail loudly
                        # rather than silently append spurious levels.
                        levels = list(planner_levels[col])
                        unknown = [v for v in observed if v not in levels]
                        if unknown:
                            return json.dumps({
                                "status": "error",
                                "message": (
                                    f"Input column '{col}' has values {unknown} that are not "
                                    f"in the planner-declared level universe {levels}. Either "
                                    f"re-encode the data to use the declared level names, or "
                                    f"correct the planner's optimization_params.levels for "
                                    f"this parameter."
                                ),
                            })
                    else:
                        levels = observed
                    level_maps[col] = levels
                    optimization_inputs.append(col)
                else:
                    if not pd.api.types.is_numeric_dtype(df[col]):
                        return json.dumps({
                            "status": "error",
                            "message": (
                                f"Input column '{col}' is non-numeric but not declared "
                                f"categorical. Either fix the data or set its input_type "
                                f"to 'categorical' in the scalarizer output."
                            ),
                            "available_columns": list(df.columns),
                        })
                    optimization_inputs.append(col)

            for w in type_conflict_warnings:
                print(f"  ⚠️ {w}")

            if not optimization_inputs:
                return json.dumps({
                    "status": "error",
                    "message": "No usable input parameters found."
                })

            self.orch.expected_input_columns = optimization_inputs
            self.orch.expected_input_levels = level_maps if level_maps else None

            # Build bounds (continuous: scientific or data-derived; categorical: [0, n-1])
            input_bounds = []
            for col in optimization_inputs:
                if col in level_maps:
                    n = len(level_maps[col])
                    input_bounds.append([0.0, float(n - 1)])
                    print(f"     -> Bound for '{col}': [0, {n - 1}] (Source: CATEGORICAL — {n} levels)")
                elif col in scientific_bounds:
                    sci_min, sci_max = scientific_bounds[col]
                    input_bounds.append([sci_min, sci_max])
                    print(f"     -> Bound for '{col}': [{sci_min}, {sci_max}] (Source: PLANNER)")
                else:
                    data_min = float(df[col].min())
                    data_max = float(df[col].max())

                    if data_min == data_max:
                        margin = 1.0 if data_min == 0 else abs(data_min * 0.1)
                    else:
                        margin = (data_max - data_min) * 0.1

                    safe_min = data_min - margin
                    safe_max = data_max + margin

                    input_bounds.append([safe_min, safe_max])
                    print(f"     -> Bound for '{col}': [{safe_min:.2f}, {safe_max:.2f}] (Source: DATA)")

            # Build cat_dims (positional indices) and integer-encode CSV
            cat_dims = [
                i for i, c in enumerate(optimization_inputs) if c in level_maps
            ]
            bo_data_path_for_run = str(self.orch.bo_data_path)
            if level_maps:
                df_encoded = df.copy()
                for col, levels in level_maps.items():
                    idx = {lv: i for i, lv in enumerate(levels)}
                    encoded = df[col].astype(str).map(idx)
                    if encoded.isnull().any():
                        unknown = sorted(
                            df.loc[encoded.isnull(), col].astype(str).unique().tolist()
                        )
                        return json.dumps({
                            "status": "error",
                            "message": (
                                f"Unknown levels for categorical input '{col}': {unknown}. "
                                f"Known levels: {levels}."
                            )
                        })
                    df_encoded[col] = encoded.astype(float)
                encoded_dir = self.orch.base_dir / "bo_artifacts"
                encoded_dir.mkdir(exist_ok=True, parents=True)
                encoded_path = encoded_dir / "optimization_data_encoded.csv"
                df_encoded.to_csv(encoded_path, index=False)
                bo_data_path_for_run = str(encoded_path)
                print(f"  🔡 Encoded {len(level_maps)} categorical column(s) → {encoded_path.name}")
            
            # ============================================
            # BATCH SIZE DETERMINATION
            # ============================================
            if not parallel_capable:
                final_batch_size = 1
                mode_desc = "sequential (single experiment)"
            else:
                if batch_size is None:
                    return json.dumps({
                        "status": "batch_size_required",
                        "message": "Batch size must be specified for parallel optimization.",
                        "instruction": (
                            "Analyze the experimental plan to determine appropriate batch_size "
                            "(e.g., plate format, number of conditions, equipment capacity), "
                            "then call: run_optimization(parallel_capable=True, batch_size=N)"
                        ),
                        "hint": "Common values: 8, 12, 24, 96, 384 for plate-based experiments"
                    })
                
                if batch_size < 1:
                    return json.dumps({
                        "status": "error", 
                        "message": f"Invalid batch_size: {batch_size}. Must be at least 1."
                    })
                
                final_batch_size = batch_size
                mode_desc = f"parallel (batch of {batch_size})"
                print(f"    ℹ️  Using batch_size: {batch_size}")
            
            # ============================================
            # CONSTRAINT-AWARE & BUDGET-AWARE LOGGING
            # ============================================
            if physical_constraints:
                mode_desc += " + constraint-aware"
                print(f"    📐 Physical constraints provided — will use LLM-guided batch design")
            
            if experimental_budget is not None:
                mode_desc += f" + budget={experimental_budget}"
                print(f"    💰 Experimental budget: {experimental_budget} iteration(s) remaining")
            
            print(f"    📊 Optimization Setup:")
            print(f"       Mode: {mode_desc}")
            print(f"       Data points: {len(df)}")
            print(f"       Inputs: {self.orch.expected_input_columns}")
            print(f"       Targets: {self.orch.expected_target_columns}")
            print(f"       Bounds: {input_bounds}")
            if physical_constraints:
                print(f"       Constraints: {physical_constraints[:100]}...")
            
            # ============================================
            # DATA SUFFICIENCY CHECK FOR MOO
            # ============================================
            n_targets = len(self.orch.expected_target_columns)
            n_inputs = len(self.orch.expected_input_columns)
            n_data = len(df)
            if n_targets > 1:
                min_recommended = 5 * n_inputs * n_targets
                if n_data < min_recommended:
                    print(f"    ⚠️  MOO data sufficiency: {n_data} points for "
                          f"{n_inputs} inputs × {n_targets} targets (recommend ≥{min_recommended})")
                    return json.dumps({
                        "status": "warning",
                        "message": (
                            f"Multi-objective optimization with {n_targets} targets and "
                            f"{n_inputs} inputs ideally needs ≥{min_recommended} data points, "
                            f"but only {n_data} are available. The Pareto recommendations "
                            f"will be unreliable."
                        ),
                        "suggestion": (
                            f"Re-call run_optimization with targets=[\"chosen_target\"] to "
                            f"narrow to single-objective. No need to re-analyze files. "
                            f"You can switch to MOO once more data is collected."
                        ),
                        "current_targets": self.orch.expected_target_columns,
                        "data_points": n_data,
                        "recommended_minimum": min_recommended,
                    })

            try:
                # ============================================
                # DISTILL OBJECTIVE & CALL BO
                # ============================================
                bo_objective = self._distill_objective_for_bo(
                    self.orch.expected_target_columns
                )
                res = self.orch.bo.run_optimization_loop(
                    data_path=bo_data_path_for_run,
                    objective_text=bo_objective,
                    input_cols=self.orch.expected_input_columns,
                    input_bounds=input_bounds,
                    target_cols=self.orch.expected_target_columns,
                    target_directions=self.orch.target_directions,
                    output_dir=str(self.orch.base_dir / "bo_artifacts"),
                    batch_size=int(final_batch_size),
                    physical_constraints=physical_constraints,
                    experimental_budget=experimental_budget,
                    strategy_hint=strategy_hint,
                    plot_acq=True,
                    save_acq=True,
                    cat_dims=cat_dims if cat_dims else None,
                )
                
                if res.get("status") != "success":
                    return json.dumps({
                        "status": "error", 
                        "message": res.get("error", "Optimization failed"),
                        "bo_output": res
                    })
                
                # Format response
                next_params = res.get('next_parameters')

                # Decode categorical recommendations back to human-readable levels
                if level_maps and next_params is not None:
                    next_params = self._decode_categorical_recs(next_params, level_maps)

                if parallel_capable:
                    hint = f"Run all {final_batch_size} experiments in parallel, then use analyze_file on each result file."
                    params_summary = f"Generated {final_batch_size} parameter sets"
                else:
                    hint = "Run this experiment, then use analyze_file on the result to continue."
                    params_summary = "Generated next experiment parameters"
                
                response = {
                    "status": "success",
                    "mode": "parallel" if parallel_capable else "sequential",
                    "batch_size": final_batch_size,
                    "recommended_parameters": next_params,
                    "params_summary": params_summary,
                    "strategy_used": res.get('strategy', {}).get('acquisition_strategy', {}).get('type'),
                    "plot_path": res.get('plot_path'),
                    "hint": hint
                }
                if res.get("acq_plot_path"):
                    response["acq_plot_path"] = res["acq_plot_path"]
                if res.get("acq_data_path"):
                    response["acq_data_path"] = res["acq_data_path"]
                
                # Include visual inspection results
                if res.get("inspection"):
                    response["inspection"] = res["inspection"]

                # Include constrained planning metadata
                if res.get("constrained_planning"):
                    cp = res["constrained_planning"]
                    response["constraint_aware"] = True
                    response["coverage_summary"] = cp.get("coverage_summary", "")
                    response["trade_offs"] = cp.get("trade_offs", "")
                    if cp.get("validation_errors"):
                        response["constraint_warnings"] = cp["validation_errors"]
                
                # Include budget context
                if res.get("budget"):
                    response["budget"] = res["budget"]
                
                return json.dumps(response)
                
            except Exception as e:
                logging.error(f"Optimization error: {e}")
                return json.dumps({
                    "status": "error",
                    "message": str(e)
                })
        
        self._register_tool(
            func=run_optimization,
            name="run_optimization",
            description=(
                "Runs Bayesian Optimization to suggest next experimental parameters. "
                "Requires at least 3 data points from analyze_file. "
                "For parallel mode, batch_size must be specified. "
                "Supports optional physical_constraints for constraint-aware batch design — "
                "when provided, the agent evaluates the acquisition landscape and uses LLM "
                "reasoning to design a batch that maximizes information gain while respecting "
                "physical experimental limitations (e.g., plate layouts, discrete reagent stocks, "
                "shared equipment channels). "
                "Supports optional experimental_budget for exploration/exploitation control — "
                "pass the number of remaining optimization iterations to shift strategy from "
                "exploration (high budget) to exploitation (low budget)."
            ),
            parameters={
                "parallel_capable": {
                    "type": "boolean",
                    "description": "True if experiments can run in parallel. False for sequential (default)."
                },
                "batch_size": {
                    "type": "integer",
                    "description": (
                        "Number of parallel experiments (required if parallel_capable=True). "
                        "Infer from experimental plan (e.g., plate format, grid size, equipment capacity)."
                    )
                },
                "physical_constraints": {
                    "type": "string",
                    "description": (
                        "Natural language description of physical experimental constraints that "
                        "prevent arbitrary parameter combinations. When provided, the optimizer "
                        "evaluates the full acquisition landscape and uses LLM reasoning to design "
                        "a realizable batch. Examples:\n"
                        "- '96-well plate: rows share temperature (8 values), columns share pH (12 values)'\n"
                        "- 'Only 5 catalyst concentrations available: 0.1, 0.5, 1.0, 2.0, 5.0 mM'\n"
                        "- 'Reactor has 4 zones with independent temp but shared pressure'\n"
                        "- 'Gradient limited to linear ramp: min at well A1, max at well H12'\n"
                        "If not provided, standard unconstrained BO is used."
                    )
                },
                "experimental_budget": {
                    "type": "integer",
                    "description": (
                        "Number of remaining optimization iterations (including this one). "
                        "Controls exploration-vs-exploitation balance:\n"
                        "- 1 = final shot (pure exploitation, no exploration)\n"
                        "- 2-3 = critical budget (strongly favor exploitation)\n"
                        "- Higher values = scaled based on campaign progress\n"
                        "- Omit for no budget constraint (default behavior).\n"
                        "Pass when user mentions remaining experiments, budget, 'last round', "
                        "or 'N more tries'. This counts iterations (calls to run_optimization), "
                        "not individual experiments within a batch."
                    )
                },
                "targets": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": (
                        "Subset of target columns to optimize. Use to narrow from "
                        "multi-objective to single-objective without re-analyzing files. "
                        "Example: targets=['Peak_Area'] to optimize only peak area when "
                        "the scalarizer extracted multiple metrics. The specified targets "
                        "must exist in the optimization data."
                    )
                },
                "strategy_hint": {
                    "type": "string",
                    "description": (
                        "Optional user preference for BO strategy. Pass when the user "
                        "requests a specific kernel, acquisition function, or noise prior. "
                        "Examples: 'use RBF kernel', 'try Thompson sampling', "
                        "'switch to Matern-1.5', 'use UCB with high exploration'. "
                        "The hint is respected unless it conflicts with budget constraints."
                    )
                }
            },
            required=[]
        )


        # 9. SAVE FILE
        def save_file(filename: str, content: str, subfolder: str = ""):
            """
            Save text content (code, protocols, notes) to a file in the session
            directory.
            """
            print(f"  ⚡ Tool: Saving file '{filename}'...")

            # Sanitise: strip path separators from filename to prevent traversal.
            safe_name = Path(filename).name
            if not safe_name:
                return json.dumps({
                    "status": "error",
                    "message": "Invalid filename.",
                })

            target_dir = self.orch.base_dir
            if subfolder:
                safe_sub = Path(subfolder).name
                target_dir = target_dir / safe_sub
            target_dir.mkdir(parents=True, exist_ok=True)
            dest = target_dir / safe_name

            try:
                dest.write_text(content, encoding="utf-8")
                print(f"    💾 Saved: {dest}")
                return json.dumps({
                    "status": "success",
                    "path": str(dest),
                    "size_bytes": dest.stat().st_size,
                })
            except Exception as e:
                logging.error(f"save_file failed: {e}")
                return json.dumps({
                    "status": "error",
                    "message": str(e),
                })

        self._register_tool(
            func=save_file,
            name="save_file",
            description=(
                "Save text content (code, protocols, scripts, notes) to a file "
                "in the session directory. Use this to persist generated code, "
                "instrument protocols, or any text artifact."
            ),
            parameters={
                "filename": {
                    "type": "string",
                    "description": (
                        "Name of the file to create, e.g. 'extraction_protocol.py' "
                        "or 'notes.txt'."
                    ),
                },
                "content": {
                    "type": "string",
                    "description": "The text content to write to the file.",
                },
                "subfolder": {
                    "type": "string",
                    "description": (
                        "Optional subfolder within the session directory, "
                        "e.g. 'protocols' or 'scripts'. Created if it doesn't exist."
                    ),
                },
            },
            required=["filename", "content"],
        )

        # 10. SAVE CHECKPOINT
        def save_checkpoint():
            """
            Saves complete orchestrator state including conversation and agent state.
            Use this periodically during long campaigns.
            """
            checkpoint_path = self.orch.base_dir / "checkpoint.json"
            
            # Calculate data points
            data_points = 0
            if self.orch.bo_data_path.exists():
                try:
                    df = pd.read_csv(self.orch.bo_data_path)
                    data_points = len(df)
                except:
                    pass
            
            # Get message count (handle both OpenAI and Gemini)
            if self.orch.use_openai:
                # OpenAI: messages is a list attribute
                message_count = len(self.orch.messages)
            else:
                # Gemini: history is in chat_session
                try:
                    message_count = len(self.orch.chat_session.history) if hasattr(self.orch.chat_session, 'history') else 0
                except:
                    message_count = 0
            
            state = {
                "timestamp": datetime.now().isoformat(),
                "objective": self.orch.objective,
                "active_scalarizer_script": self.orch.active_scalarizer_script,
                "expected_input_columns": self.orch.expected_input_columns,
                "expected_target_columns": self.orch.expected_target_columns,
                "data_points_collected": data_points,
                "message_count": message_count,
                "planner_state": self.orch.planner.state if hasattr(self.orch.planner, 'state') else None,
                "latest_tea_results": self.orch.latest_tea_results,
                "autonomy_level": self.orch.autonomy_level.value if hasattr(self.orch, 'autonomy_level') and self.orch.autonomy_level else None,
                "data_dir": str(self.orch.data_dir) if self.orch.data_dir else None,
                "knowledge_dir": str(self.orch.knowledge_dir) if self.orch.knowledge_dir else None,
                "code_dir": str(self.orch.code_dir) if self.orch.code_dir else None,
            }
            
            try:
                with open(checkpoint_path, 'w') as f:
                    json.dump(state, f, indent=2)
                
                print(f"    💾 Checkpoint saved: {checkpoint_path}")
                
                result = {
                    "status": "success",
                    "checkpoint_path": str(checkpoint_path),
                    "data_points": data_points,
                    "message_count": message_count,
                    "timestamp": state["timestamp"]
                }

                # Check if knowledge synthesis might be valuable
                planner_state = self.orch.planner.state if self.orch.planner.state else {}
                plan_history = planner_state.get("plan_history", [])
                iterations_with_results = len(planner_state.get("experimental_results", []))
                existing_knowledge = len(self.orch.active_knowledge)

                if iterations_with_results >= 2 and existing_knowledge == 0:
                    result["knowledge_synthesis_available"] = True
                    result["plan_iterations_with_results"] = iterations_with_results
                elif existing_knowledge > 0 and iterations_with_results > existing_knowledge:
                    result["knowledge_update_available"] = True
                    result["unsynthesized_iterations"] = iterations_with_results - existing_knowledge

                return json.dumps(result)
                
            except Exception as e:
                logging.error(f"Checkpoint save failed: {e}")
                return json.dumps({
                    "status": "error",
                    "message": f"Failed to save checkpoint: {e}"
                })
        
        # Register the tool
        self._register_tool(
            func=save_checkpoint,
            name="save_checkpoint",
            description=(
                "Saves complete campaign state including conversation history, "
                "analysis scripts, and optimization data. Use this periodically "
                "during long campaigns (every 3-5 experiments) to enable resumption "
                "after crashes or breaks."
            ),
            parameters={},
            required=[]
        )

        # 10. DISCARD PLAN
        def discard_plan(reason: str = ""):
            """
            Discards the most recent experimental plan (marks it as superseded).
            The plan remains in history for transparency but won't appear in reports.
            
            Args:
                reason: Why the plan is being discarded
            """
            if not self.orch.planner.state:
                return json.dumps({
                    "status": "error",
                    "message": "No active planning session"
                })
            
            history = self.orch.planner.state.get("plan_history", [])
            
            if not history:
                return json.dumps({
                    "status": "error",
                    "message": "No plans in history to discard"
                })
            
            # Find last non-TEA, non-superseded entry
            for i in range(len(history) - 1, -1, -1):
                plan = history[i]
                if (plan.get("type") != "technoeconomic_analysis" and 
                    plan.get("status") != "superseded"):
                    
                    # Mark as superseded instead of deleting
                    plan["status"] = "superseded"
                    plan["superseded_reason"] = reason if reason else "Plan replaced with corrected version"
                    plan["superseded_at"] = datetime.now().isoformat()
                    
                    print(f"    🗑️  Discarded plan: iteration {plan.get('iteration')}")
                    if reason:
                        print(f"       Reason: {reason}")
                    
                    return json.dumps({
                        "status": "success",
                        "message": f"Plan from iteration {plan.get('iteration')} discarded",
                        "reason": plan["superseded_reason"],
                        "hint": "The discarded plan remains in history for transparency"
                    })
            
            return json.dumps({
                "status": "error",
                "message": "No active experimental plans to discard"
            })

        # Register the tool
        self._register_tool(
            func=discard_plan,
            name="discard_plan",
            description=(
                "Discards the most recent experimental plan (marks it as superseded). "
                "The plan remains in full history for transparency but won't appear in final reports. "
                "Use when correcting a wrong plan before generating the corrected version."
            ),
            parameters={
                "reason": {
                    "type": "string",
                    "description": (
                        "Why the plan is being discarded. Be specific about the mismatch. "
                        "Examples: 'Wrong material - data has Mg not Mn', "
                        "'User requested different equipment', 'Incorrect objective interpretation'"
                    )
                }
            },
            required=["reason"]
        )

        def show_directory_guide():
            """
            Shows the recommended directory structure for optimal agent performance.
            """
            guide = """
        ╔══════════════════════════════════════════════════════════════════════════╗
        ║                  RECOMMENDED DIRECTORY STRUCTURE                         ║
        ╚══════════════════════════════════════════════════════════════════════════╝

        📁 my_research_project/          ← Run orchestrator from here
        │
        ├── 📚 papers/                    ← Scientific papers & literature
        │   ├── separation_methods_2024.pdf
        │   ├── lithium_extraction_review.pdf
        │   └── rare_earth_recovery.pdf
        │
        ├── 📊 experimental_results/      ← Raw experimental data files
        │   ├── batch_001.csv
        │   ├── batch_002.csv
        │   ├── batch_003.csv
        │   └── pilot_run_*.xlsx
        │
        ├── 💻 code/                      ← Analysis scripts & API docs (optional)
        │   ├── analysis_pipeline.py
        │   ├── visualization.py
        │   └── api_documentation/
        │
        ├── 📁 campaign_session/          ← Created automatically by orchestrator
        │   ├── optimization_data.csv    (collected metrics)
        │   ├── analysis_artifacts/      (generated analysis scripts)
        │   ├── bo_artifacts/            (optimization plots)
        │   ├── plan.json                (experimental plans)
        │   └── checkpoint.json          (saved state)
        │
        └── 🗂️ kb_storage/                ← Created automatically
            ├── default_kb_docs/         (knowledge base from papers)
            └── default_kb_code/         (knowledge base from code)

        ╔══════════════════════════════════════════════════════════════════════════╗
        ║                           QUICK START GUIDE                              ║
        ╚══════════════════════════════════════════════════════════════════════════╝

        CHAT EXAMPLES:

        📋 Generate plan with papers:
        "Generate a plan for lithium extraction using ./papers/ and ./code/"

        📊 Analyze experimental data:
        "Analyze ./experimental_results/batch_001.csv and extract yield"

        🔬 Run optimization:
        "Run optimization to suggest next experiments"

        💾 Save progress:
        "Save checkpoint"
        """
            
            print(guide)
            
            # Also return as JSON for the LLM
            return json.dumps({
                "status": "success",
                "message": "Directory structure guide displayed",
                "recommended_folders": ["papers/", "experimental_results/", "code/"],
                "auto_created_folders": ["campaign_session/", "kb_storage/"]
            })

        # Register the tool
        self._register_tool(
            func=show_directory_guide,
            name="show_directory_guide",
            description=(
                "Shows recommended directory structure for optimal agent performance. "
                "Use when user asks about setup, organization, or how to structure their project."
            ),
            parameters={},
            required=[]
        )

        # =====================================================================
        # READ FILE (non-destructive inspection)
        # =====================================================================
        def read_file(file_path: str, max_lines: int = 200) -> str:
            """
            Read and return the contents of a file. Use this to inspect
            plans, protocols, configs, logs, or any text/JSON file without
            triggering analysis pipelines.
            """
            print(f"  ⚡ Tool: Reading file '{file_path}'...")

            # Resolve path
            resolved, error = self._resolve_data_path(file_path)
            if error:
                return error

            path = Path(resolved)
            if not path.is_file():
                return json.dumps({
                    "status": "error",
                    "message": f"Not a file: {file_path}"
                })

            try:
                ext = path.suffix.lower()

                # Size guard — skip for Excel/CSV since we cap at 50 rows × 40 cols
                if ext not in ('.xlsx', '.xls', '.csv'):
                    size_mb = path.stat().st_size / (1024 * 1024)
                    if size_mb > 5:
                        return json.dumps({
                            "status": "error",
                            "message": f"File too large ({size_mb:.1f} MB)."
                        })

                if ext == ".json":
                    with open(path, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                    content = json.dumps(data, indent=2)
                elif ext in ('.xlsx', '.xls', '.csv'):
                    MAX_PREVIEW_ROWS = 100
                    MAX_PREVIEW_COLS = 40
                    MAX_PREVIEW_CHARS = 30000
                    if ext == '.csv':
                        df_preview = pd.read_csv(path, nrows=MAX_PREVIEW_ROWS)
                        with open(path) as _f:
                            total_rows = sum(1 for _ in _f) - 1
                    else:
                        df_preview = pd.read_excel(path, nrows=MAX_PREVIEW_ROWS)
                        try:
                            import openpyxl
                            _wb = openpyxl.load_workbook(path, read_only=True)
                            total_rows = _wb.active.max_row - 1
                            _wb.close()
                        except Exception:
                            total_rows = len(df_preview)
                    total_cols = len(df_preview.columns)
                    display_df = df_preview.iloc[:, :MAX_PREVIEW_COLS]
                    preview_text = display_df.to_string()
                    # Adaptive row reduction if output exceeds char budget
                    if len(preview_text) > MAX_PREVIEW_CHARS and len(display_df) > 5:
                        ratio = MAX_PREVIEW_CHARS / len(preview_text)
                        fewer_rows = max(5, int(len(display_df) * ratio))
                        display_df = display_df.iloc[:fewer_rows]
                        preview_text = display_df.to_string()
                        if len(preview_text) > MAX_PREVIEW_CHARS:
                            preview_text = preview_text[:MAX_PREVIEW_CHARS] + "\n... (truncated)"
                    shown_rows = len(display_df)
                    shown_cols = len(display_df.columns)
                    trunc_parts = []
                    if shown_rows < total_rows:
                        trunc_parts.append(f"first {shown_rows} rows")
                    if shown_cols < total_cols:
                        trunc_parts.append(f"first {shown_cols} columns")
                    trunc = f" (showing {', '.join(trunc_parts)})" if trunc_parts else ""
                    content = f"Shape: {total_rows} rows × {total_cols} columns{trunc}\n\n{preview_text}"
                else:
                    with open(path, 'r', encoding='utf-8', errors='replace') as f:
                        lines = f.readlines()
                    if len(lines) > max_lines:
                        content = "".join(lines[:max_lines])
                        content += f"\n... ({len(lines) - max_lines} more lines truncated)"
                    else:
                        content = "".join(lines)

                return json.dumps({
                    "status": "success",
                    "file_path": str(path),
                    "content": content
                })

            except Exception as e:
                return json.dumps({
                    "status": "error",
                    "message": f"Failed to read file: {e}"
                })

        self._register_tool(
            func=read_file,
            name="read_file",
            description=(
                "Read and return the contents of a text or JSON file. "
                "Use this to inspect plans, protocols, scripts, configs, or logs. "
                "Do NOT use analyze_file for reading — that triggers the scalarizer pipeline."
            ),
            parameters={
                "file_path": {
                    "type": "string",
                    "description": "Path to the file to read"
                },
                "max_lines": {
                    "type": "integer",
                    "description": "Maximum lines to return for large files (default: 200)"
                }
            },
            required=["file_path"]
        )

        # =====================================================================
        # KNOWLEDGE DATA QUERY TOOL
        # =====================================================================

        QUERYABLE_EXTENSIONS = {'.xlsx', '.xls', '.csv'}

        DIR_DB_MIN_FILES = 10  # minimum same-extension files to treat as database

        def _summarize_json_value(value, max_str_len=200):
            """Return a compact string representation of a JSON value."""
            if value is None:
                return "null"
            if isinstance(value, bool):
                return str(value).lower()
            if isinstance(value, (int, float)):
                return str(value)
            if isinstance(value, str):
                if len(value) > max_str_len:
                    return json.dumps(value[:100]) + f"... ({len(value)} chars)"
                return json.dumps(value)
            if isinstance(value, list):
                if len(value) == 0:
                    return "list (0 items)"
                first = _summarize_json_value(value[0], max_str_len=80)
                return f"list ({len(value)} items, first: {first})"
            if isinstance(value, dict):
                keys = list(value.keys())
                return f"dict ({len(keys)} keys: {keys[:5]})"
            return repr(value)[:100]

        def _summarize_json_file(file_path: str) -> str:
            """Parse a JSON file and return a compact summary showing all keys."""
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            if isinstance(data, dict):
                lines = ["{"]
                for k, v in data.items():
                    lines.append(f"  {json.dumps(k)}: {_summarize_json_value(v)},")
                lines.append("}")
                return "\n".join(lines)
            elif isinstance(data, list):
                if len(data) == 0:
                    return "[]  (empty list)"
                summary = f"list of {len(data)} items\n"
                if isinstance(data[0], dict):
                    summary += "First item:\n"
                    lines = ["{"]
                    for k, v in data[0].items():
                        lines.append(f"  {json.dumps(k)}: {_summarize_json_value(v)},")
                    lines.append("}")
                    summary += "\n".join(lines)
                else:
                    summary += f"First item: {_summarize_json_value(data[0])}"
                return summary
            return repr(data)[:2000]

        def _inspect_directory(dir_path: str) -> dict:
            """Summarize directory contents for LLM-driven querying."""
            from collections import Counter

            p = Path(dir_path)
            files_by_ext = Counter()
            all_files = {}  # ext -> sorted list of filenames

            for f in p.iterdir():
                if f.is_file() and not f.name.startswith('.'):
                    ext = f.suffix.lower()
                    files_by_ext[ext] += 1
                    all_files.setdefault(ext, []).append(f.name)

            # Sort filenames for each extension
            for ext in all_files:
                all_files[ext].sort()

            total = sum(files_by_ext.values())

            # Pick dominant extension for sampling
            if not files_by_ext:
                return {
                    "directory": str(p),
                    "files_by_extension": {},
                    "total_files": 0,
                    "sample_files": [],
                    "all_filenames_sample": [],
                }

            # Sample files from each extension that has >= DIR_DB_MIN_FILES
            # For smaller groups, sample one file; this gives the LLM visibility
            # into all file types in the directory.
            sample_files = []  # list of {"name", "content", "ext"}
            sampled_exts = []
            for ext, count in files_by_ext.most_common():
                if count < DIR_DB_MIN_FILES:
                    continue
                sampled_exts.append(ext)
                names_for_ext = all_files[ext]
                # Pick one representative file per extension
                sample_name = names_for_ext[len(names_for_ext) // 2]
                fp = p / sample_name
                try:
                    if ext == '.json':
                        content = _summarize_json_file(str(fp))
                    else:
                        size = fp.stat().st_size
                        with open(fp, 'r', encoding='utf-8', errors='replace') as fh:
                            if size > 50_000:
                                content = fh.read(50_000)
                                content += f"\n... (truncated, {size} bytes total)"
                            else:
                                content = fh.read()
                except Exception as e:
                    content = f"(error reading file: {e})"
                sample_files.append({"name": sample_name, "content": content, "ext": ext})

            # Collect first 20 filenames from the most common extension
            dominant_ext = files_by_ext.most_common(1)[0][0]

            return {
                "directory": str(p),
                "files_by_extension": dict(files_by_ext.most_common()),
                "total_files": total,
                "extensions": sampled_exts,
                "sample_files": sample_files,
                "all_filenames_sample": all_files[dominant_ext][:20],
            }

        def _inspect_knowledge_file(file_path: str) -> dict:
            """Return a format-agnostic diagnostic snapshot of a data file."""
            p = Path(file_path)
            ext = p.suffix.lower()

            if ext in ('.xlsx', '.xls'):
                df = pd.read_excel(file_path, nrows=10)
                read_instr = f"pd.read_excel('{file_path}')"
                fmt = "excel"
            elif ext == '.csv':
                df = pd.read_csv(file_path, nrows=10)
                read_instr = f"pd.read_csv('{file_path}')"
                fmt = "csv"
            else:
                return {"error": f"Unsupported format: {ext}"}

            # Get actual row count without loading full file
            if ext in ('.xlsx', '.xls'):
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(file_path, read_only=True)
                    total_rows = wb.active.max_row - 1
                    wb.close()
                except Exception:
                    total_rows = "unknown"
            else:
                with open(file_path) as f:
                    total_rows = sum(1 for _ in f) - 1

            return {
                "format": fmt,
                "shape": (total_rows, len(df.columns)),
                "columns": list(df.columns),
                "dtypes": df.dtypes.to_string(),
                "head": df.to_string(),
                "read_instruction": read_instr,
            }

        def _discover_queryable_files() -> list:
            """Scan knowledge directories for queryable data files and directory databases."""
            from collections import Counter

            search_dirs = set()
            # User-specified knowledge folder (or kb_storage/ default)
            if self.orch.knowledge_dir and Path(self.orch.knowledge_dir).exists():
                search_dirs.add(Path(self.orch.knowledge_dir))
            # Session knowledge dir (where UI uploads go)
            session_kdir = Path(self.orch.base_dir) / "knowledge"
            if session_kdir.exists():
                search_dirs.add(session_kdir)

            if not search_dirs:
                return []

            found = {}
            dir_db_dirs = set()  # directories detected as databases

            # --- Detect directory databases ---
            # Check each search dir and its immediate subdirectories
            dirs_to_check = list(search_dirs)
            for kdir in search_dirs:
                for child in kdir.iterdir():
                    if child.is_dir() and not child.name.startswith('.'):
                        dirs_to_check.append(child)

            for d in dirs_to_check:
                ext_counts = Counter()
                for f in d.iterdir():
                    if f.is_file() and not f.name.startswith('.'):
                        ext_counts[f.suffix.lower()] += 1
                # If any extension has enough files, treat whole directory as database
                db_exts = {ext: count for ext, count in ext_counts.items()
                           if count >= DIR_DB_MIN_FILES}
                if db_exts:
                    parts = ", ".join(f"{c} {e}" for e, c in sorted(db_exts.items()))
                    display_name = f"{d.name} ({parts} files)"
                    found[display_name] = {
                        "name": display_name,
                        "path": str(d),
                        "type": "directory",
                    }
                    dir_db_dirs.add(d)

            # --- Discover single queryable files ---
            for kdir in search_dirs:
                for f in kdir.rglob("*"):
                    if f.is_file() and f.suffix.lower() in QUERYABLE_EXTENSIONS:
                        # Skip files inside a detected database directory
                        if any(f.parent == dd for dd in dir_db_dirs):
                            continue
                        found[f.name] = {"name": f.name, "path": str(f), "type": "file"}

            return sorted(found.values(), key=lambda x: x["name"])

        def _resolve_knowledge_data_file(file_name: str):
            """Resolve a file name to a path in the knowledge directory.

            Returns (target, error) where target is either:
            - a file path string (for single files)
            - a dict with "type": "directory" (for directory databases)
            """
            from difflib import get_close_matches
            candidates = _discover_queryable_files()
            if not candidates:
                return None, json.dumps({
                    "status": "error",
                    "message": "No queryable data files or directories found in knowledge directory."
                })
            names = [c["name"] for c in candidates]
            entry_map = {c["name"]: c for c in candidates}

            def _return_entry(name):
                entry = entry_map[name]
                if entry.get("type") == "directory":
                    return entry, None
                return entry["path"], None

            # Exact match
            if file_name in entry_map:
                return _return_entry(file_name)

            # Stem match (without extension) — only for file entries
            stem = Path(file_name).stem
            for n in names:
                if Path(n).stem == stem:
                    return _return_entry(n)

            # Fuzzy match
            matches = get_close_matches(file_name, names, n=3, cutoff=0.5)
            if matches:
                suggestion = ", ".join(matches)
                return None, json.dumps({
                    "status": "error",
                    "message": f"'{file_name}' not found. Did you mean: {suggestion}?",
                    "available_files": names
                })

            return None, json.dumps({
                "status": "error",
                "message": f"'{file_name}' not found.",
                "available_files": names
            })

        def _extract_code_block(text: str) -> str:
            """Extract Python code from LLM response."""
            # Try ```python blocks
            match = re.search(r'```python\s*\n(.*?)```', text, re.DOTALL)
            if match:
                return match.group(1).strip()
            # Try generic ``` blocks
            match = re.search(r'```\s*\n(.*?)```', text, re.DOTALL)
            if match:
                return match.group(1).strip()
            # Prompt ends with open ```python — response may be code followed by ```
            match = re.search(r'^(.*?)```', text, re.DOTALL)
            if match and 'import ' in match.group(1):
                return match.group(1).strip()
            # Try raw code starting with import
            for line in text.split('\n'):
                if line.strip().startswith('import '):
                    idx = text.index(line)
                    # Strip any trailing ``` if present
                    code = text[idx:].strip()
                    code = re.sub(r'\n```\s*$', '', code)
                    return code
            # Last resort: if entire response looks like code (has import somewhere)
            if 'import ' in text and 'print(' in text:
                return re.sub(r'\n```\s*$', '', text.strip())
            return ""

        def _build_directory_scaffold(info: dict) -> str:
            """Build scaffold code with readers for each file type."""
            dir_path = info["directory"]
            lines = [
                "import json, glob, os",
                "from pathlib import Path",
                "",
                f"directory = {json.dumps(dir_path)}",
            ]
            # Add file lists and readers per extension
            reader_map = {
                '.json': (
                    "def read_json(filepath):\n"
                    "    with open(filepath, 'r') as f:\n"
                    "        return json.load(f)"
                ),
                '.csv': (
                    "import pandas as pd\n"
                    "def read_csv(filepath):\n"
                    "    return pd.read_csv(filepath)"
                ),
                '.tsv': (
                    "import pandas as pd\n"
                    "def read_tsv(filepath):\n"
                    "    return pd.read_csv(filepath, sep='\\t')"
                ),
            }
            default_reader = (
                "def read_text(filepath):\n"
                "    with open(filepath, 'r', encoding='utf-8', errors='replace') as f:\n"
                "        return f.read()"
            )

            for ext in info["extensions"]:
                count = info["files_by_extension"].get(ext, 0)
                var_name = ext.lstrip('.') + "_files"
                lines.append(f'{var_name} = sorted(glob.glob(os.path.join(directory, "*{ext}")))')
                lines.append(f"# {count} {ext} files")
            lines.append("")

            added_readers = set()
            for ext in info["extensions"]:
                if ext in reader_map and ext not in added_readers:
                    lines.append(reader_map[ext])
                    added_readers.add(ext)
                elif ext not in added_readers:
                    lines.append(default_reader)
                    added_readers.add('_text')
            lines.append("")

            return "\n".join(lines)

        def _query_directory(dir_entry: dict, query: str) -> str:
            """Query a directory database using LLM-generated code."""
            import subprocess

            dir_path = dir_entry["path"]
            print(f"    - Inspecting directory: {dir_entry['name']}")

            try:
                info = _inspect_directory(dir_path)
            except Exception as e:
                return json.dumps({"status": "error", "message": f"Failed to inspect directory: {e}"})

            if not info["sample_files"]:
                return json.dumps({"status": "error", "message": "Directory is empty or unreadable."})

            # Build scaffold and sample sections for prompt
            scaffold = _build_directory_scaffold(info)
            sample_sections = []
            for s in info["sample_files"]:
                sample_sections.append(f"--- {s['name']} ({s['ext']}) ---\n{s['content']}")
            sample_text = "\nSAMPLE FILE CONTENTS:\n" + "\n\n".join(sample_sections) if sample_sections else ""

            prompt = KNOWLEDGE_QUERY_DIRECTORY_CODEGEN_PROMPT.format(
                directory=info["directory"],
                files_by_extension=info["files_by_extension"],
                total_files=info["total_files"],
                filenames=info["all_filenames_sample"],
                sample_sections=sample_text,
                scaffold=scaffold,
                query=query,
            )

            scripts_dir = Path(self.orch.base_dir) / "knowledge_query_scripts"
            scripts_dir.mkdir(parents=True, exist_ok=True)

            last_error = None
            for attempt in range(2):
                current_prompt = prompt
                if last_error:
                    current_prompt += f"\n\n**PREVIOUS ERROR:** {last_error}\nFix the script."

                try:
                    from .parser_utils import parse_json_from_response
                    response = self.orch.planner.model.generate_content(
                        [current_prompt],
                        generation_config={"max_output_tokens": 8192, "temperature": 0.0},
                    )
                    # Log raw response for debugging
                    raw_log_path = scripts_dir / f"kq_dir_raw_{abs(hash(query)) % 10000:04d}_a{attempt}.txt"
                    raw_log_path.write_text(response.text)
                    # LLM returns JSON: {"code": "...TODO lines..."}
                    result, parse_error = parse_json_from_response(response)
                    if parse_error or not result or "code" not in result:
                        # Fallback: treat response as raw code
                        body = _extract_code_block(response.text) or response.text.strip()
                        body = re.sub(r'^```\w*\s*', '', body)
                        body = re.sub(r'\s*```$', '', body)
                    else:
                        body = result["code"]
                    code = (
                        f"{scaffold}\n"
                        f"{body}\n"
                        f"print(json.dumps({{\"answer\": answer, \"summary\": summary}}))\n"
                    )
                except Exception as e:
                    return json.dumps({"status": "error", "message": f"Code generation failed: {e}"})

                script_path = scripts_dir / f"kq_dir_{Path(dir_path).name}_{abs(hash(query)) % 10000:04d}.py"
                script_path.write_text(code)
                print(f"    - Running: {script_path.name} (attempt {attempt + 1})")

                try:
                    result = subprocess.run(
                        ["python", str(script_path)],
                        capture_output=True, text=True, timeout=120,
                    )
                    if result.returncode != 0:
                        last_error = result.stderr.strip()[-500:]
                        continue

                    # Parse the last valid JSON object from stdout
                    # (LLM may print extra output before the json.dumps line)
                    answer_data = None
                    for line in reversed(result.stdout.strip().splitlines()):
                        line = line.strip()
                        if line.startswith('{') and line.endswith('}'):
                            try:
                                answer_data = json.loads(line)
                                if "answer" in answer_data or "summary" in answer_data:
                                    break
                            except json.JSONDecodeError:
                                continue
                    if not answer_data:
                        last_error = f"No valid JSON in output: {result.stdout[-300:]}"
                        continue

                    answer_str = json.dumps(answer_data.get("answer", ""))
                    if len(answer_str) > 5000:
                        answer_data["answer"] = str(answer_data["answer"])[:5000] + "... (truncated)"

                    print(f"    - ✅ Directory query answered successfully.")
                    return json.dumps({
                        "status": "success",
                        "query": query,
                        "file": dir_entry["name"],
                        "answer": answer_data.get("answer"),
                        "summary": answer_data.get("summary", ""),
                        "details": answer_data.get("details"),
                        "script_path": str(script_path),
                    })

                except subprocess.TimeoutExpired:
                    last_error = "Script timed out (120s limit)."
                    continue
                except json.JSONDecodeError as e:
                    last_error = f"Invalid JSON in output: {e}"
                    continue

            return json.dumps({
                "status": "error",
                "message": "Directory query failed after 2 attempts.",
                "last_error": last_error,
            })

        def query_knowledge_data(query: str, file_name: str = None) -> str:
            """Query a knowledge data file or directory database with natural language."""
            import subprocess

            print(f"  ⚡ Tool: Querying knowledge data: '{query[:80]}...'")

            # 1. Discover queryable files and directory databases
            queryable = _discover_queryable_files()
            if not queryable:
                return json.dumps({
                    "status": "error",
                    "message": "No queryable data files or directories found in knowledge directory."
                })

            # 2. Resolve target
            if file_name is None:
                if len(queryable) == 1:
                    target = queryable[0]
                    print(f"    - Auto-selected: {target['name']}")
                else:
                    return json.dumps({
                        "status": "file_selection_needed",
                        "message": "Multiple queryable sources found. Specify file_name.",
                        "available_files": [f["name"] for f in queryable]
                    })
            else:
                target, error = _resolve_knowledge_data_file(file_name)
                if error:
                    return error

            # 2b. Branch: directory database vs single file
            if isinstance(target, dict) and target.get("type") == "directory":
                return _query_directory(target, query)

            # Single file path — extract path string
            target_path = target if isinstance(target, str) else target["path"]

            # 3. Inspect file
            try:
                info = _inspect_knowledge_file(target_path)
                if "error" in info:
                    return json.dumps({"status": "error", "message": info["error"]})
            except Exception as e:
                return json.dumps({"status": "error", "message": f"Failed to read file: {e}"})

            # 4. Build prompt
            prompt = KNOWLEDGE_QUERY_CODEGEN_PROMPT.format(
                file_path=target_path,
                file_format=info["format"],
                rows=info["shape"][0],
                cols=info["shape"][1],
                columns=info["columns"],
                dtypes=info["dtypes"],
                head=info["head"],
                read_instruction=info["read_instruction"],
                query=query,
            )

            # 5. Generate and execute (with 1 retry)
            scripts_dir = Path(self.orch.base_dir) / "knowledge_query_scripts"
            scripts_dir.mkdir(parents=True, exist_ok=True)

            last_error = None
            for attempt in range(2):
                current_prompt = prompt
                if last_error:
                    current_prompt += f"\n\n**PREVIOUS ERROR:** {last_error}\nFix the script."

                # Generate code
                try:
                    response = self.orch.planner.model.generate_content(
                        [current_prompt],
                        generation_config={"max_output_tokens": 1024, "temperature": 0.0},
                    )
                    # LLM returns JSON: {"code": "...TODO lines..."}
                    from .parser_utils import parse_json_from_response
                    result, parse_error = parse_json_from_response(response)
                    if parse_error or not result or "code" not in result:
                        # Fallback: treat response as raw code
                        body = _extract_code_block(response.text) or response.text.strip()
                        body = re.sub(r'^```\w*\s*', '', body)
                        body = re.sub(r'\s*```$', '', body)
                    else:
                        body = result["code"]
                    code = (
                        f"import pandas as pd, json\n"
                        f"df = {info['read_instruction']}\n"
                        f"{body}\n"
                        f"print(json.dumps({{\"answer\": answer, \"summary\": summary}}))\n"
                    )
                except Exception as e:
                    return json.dumps({"status": "error", "message": f"Code generation failed: {e}"})

                # Write and execute script
                script_path = scripts_dir / f"kq_{Path(target_path).stem}_{abs(hash(query)) % 10000:04d}.py"
                script_path.write_text(code)
                print(f"    - Running: {script_path.name} (attempt {attempt + 1})")

                try:
                    result = subprocess.run(
                        ["python", str(script_path)],
                        capture_output=True, text=True, timeout=60,
                    )
                    if result.returncode != 0:
                        last_error = result.stderr.strip()[-500:]
                        continue

                    json_match = re.search(r'\{.*\}', result.stdout.strip(), re.DOTALL)
                    if not json_match:
                        last_error = f"No JSON in output: {result.stdout[:300]}"
                        continue

                    answer_data = json.loads(json_match.group(0))

                    # Truncate large results
                    answer_str = json.dumps(answer_data.get("answer", ""))
                    if len(answer_str) > 5000:
                        answer_data["answer"] = str(answer_data["answer"])[:5000] + "... (truncated)"

                    print(f"    - ✅ Query answered successfully.")
                    return json.dumps({
                        "status": "success",
                        "query": query,
                        "file": Path(target_path).name,
                        "answer": answer_data.get("answer"),
                        "summary": answer_data.get("summary", ""),
                        "details": answer_data.get("details"),
                        "script_path": str(script_path),
                    })

                except subprocess.TimeoutExpired:
                    last_error = "Script timed out (60s limit)."
                    continue
                except json.JSONDecodeError as e:
                    last_error = f"Invalid JSON in output: {e}"
                    continue

            return json.dumps({
                "status": "error",
                "message": f"Query failed after 2 attempts.",
                "last_error": last_error,
            })

        self._register_tool(
            func=query_knowledge_data,
            name="query_knowledge_data",
            description=(
                "Query knowledge data with natural language. Works with single "
                "data files (CSV, XLSX) and directory databases (folders of "
                "uniformly-structured files like JSON records). Generates and "
                "executes a Python script to answer questions about the data."
            ),
            parameters={
                "query": {
                    "type": "string",
                    "description": "Natural language question about the data"
                },
                "file_name": {
                    "type": "string",
                    "description": "Name of knowledge file to query (e.g., 'PWSdatabase.xlsx'). If omitted, lists available files."
                }
            },
            required=["query"]
        )

        # =====================================================================
        # KNOWLEDGE & SKILL TOOLS
        # =====================================================================

        # 12. SYNTHESIZE KNOWLEDGE
        def synthesize_knowledge(plan_ids: list, focus: str, synthesis_type: str = "reference") -> str:
            """
            Distill findings from completed planning iterations into reusable knowledge.
            The synthesized knowledge can be graduated into a skill for future campaigns.
            """
            from scilink.knowledge import synthesize_knowledge as _synthesize

            print(f"  ⚡ Tool: Synthesizing knowledge ({synthesis_type}) from {len(plan_ids)} plan iterations...")

            planner_state = self.orch.planner.state if self.orch.planner.state else {}
            plan_history = planner_state.get("plan_history", [])
            experimental_results = planner_state.get("experimental_results", [])
            feedback_history = planner_state.get("human_feedback_history", [])
            results = []
            missing_ids = []

            for pid in plan_ids:
                found = False
                # Collect ALL plan_history entries for this iteration (draft, refined, constraint-adjusted)
                matching_plans = [p for p in plan_history if str(p.get("iteration")) == str(pid)]

                if matching_plans:
                    parts = []
                    for plan in matching_plans:
                        stage = plan.get("stage", "Unknown")
                        parts.append(f"--- Stage: {stage} ---")
                        for exp in plan.get("proposed_experiments", []):
                            parts.append(f"Experiment: {exp.get('experiment_name', '')}")
                            parts.append(f"Hypothesis: {exp.get('hypothesis', '')}")
                            steps = exp.get("experimental_steps", [])
                            if steps:
                                parts.append(f"Steps: {'; '.join(steps)}")
                            parts.append(f"Justification: {exp.get('justification', '')}")
                            parts.append(f"Expected outcome: {exp.get('expected_outcome', '')}")

                    # Include experimental results/outcomes for this iteration
                    matching_results = [
                        r for r in experimental_results
                        if str(r.get("iteration")) == str(pid)
                    ]
                    for exp_result in matching_results:
                        data_summary = exp_result.get("data_summary", "")
                        if data_summary:
                            parts.append(f"--- Experimental Outcome (iteration {pid}) ---")
                            parts.append(data_summary)

                    # Collect human feedback entries relevant to this iteration
                    user_feedback_parts = []
                    for fb in feedback_history:
                        feedback_text = fb.get("feedback", "")
                        phase = fb.get("phase", "")
                        if feedback_text:
                            user_feedback_parts.append(f"[{phase}] {feedback_text}")

                    result_dict = {
                        "detailed_analysis": "\n".join(parts),
                        "analysis_id": f"plan_iter_{pid}",
                        "status": matching_plans[-1].get("stage", ""),
                    }
                    if user_feedback_parts:
                        result_dict["human_feedback"] = {
                            "user_feedback": "\n".join(user_feedback_parts)
                        }
                    results.append(result_dict)
                    found = True

                if not found:
                    missing_ids.append(pid)

            if missing_ids:
                available = sorted(set(
                    str(p.get("iteration")) for p in plan_history if p.get("iteration") is not None
                ))
                return json.dumps({
                    "status": "error",
                    "message": f"Plan iteration(s) not found: {missing_ids}",
                    "available_iterations": available
                })

            if not results:
                return json.dumps({
                    "status": "error",
                    "message": "No plan history available. Generate a plan first."
                })

            # Synthesize via the standalone function
            counter = len(self.orch.active_knowledge) + 1
            try:
                entry = _synthesize(
                    results, focus,
                    model=self.orch.planner.model,
                    knowledge_id=f"knowledge_{counter:03d}",
                    synthesis_type=synthesis_type,
                )
            except (ValueError, RuntimeError) as e:
                return json.dumps({"status": "error", "message": str(e)})

            entry["source_plans"] = plan_ids
            self.orch.active_knowledge.append(entry)

            # Save to disk
            knowledge_dir = self.orch.base_dir / "knowledge"
            knowledge_dir.mkdir(parents=True, exist_ok=True)
            knowledge_file = knowledge_dir / f"{entry['id']}.json"
            with open(knowledge_file, 'w') as f:
                json.dump(entry, f, indent=2)

            response = {
                "status": "success",
                "knowledge_id": entry["id"],
                "focus": focus,
                "synthesis_type": synthesis_type,
                "summary": entry["summary"],
                "key_findings": entry["key_findings"],
                "saved_to": str(knowledge_file),
                "note": "Use graduate_to_skill to convert this knowledge into a reusable domain skill."
            }

            # Check if any graduated skill is linked to knowledge with same focus
            for skill_name, source_ids in self.orch._graduated_skill_sources.items():
                for kid in source_ids:
                    for k in self.orch.active_knowledge:
                        if k.get("id") == kid and k.get("focus", "").lower() == focus.lower():
                            response["skill_update_suggested"] = skill_name
                            response["skill_update_note"] = (
                                f"Graduated skill '{skill_name}' is linked to knowledge "
                                f"with the same focus area. Consider calling update_skill."
                            )
                            break
                    if "skill_update_suggested" in response:
                        break
                if "skill_update_suggested" in response:
                    break

            return json.dumps(response)

        self._register_tool(
            func=synthesize_knowledge,
            name="synthesize_knowledge",
            description=(
                "Distill findings from completed planning iterations into reusable knowledge. "
                "Use when the user wants to capture learnings from plan iterations — e.g., "
                "experimental design patterns, parameter ranges that worked, or failure modes."
            ),
            parameters={
                "plan_ids": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "List of plan iteration numbers (as strings) to synthesize knowledge from"
                },
                "focus": {
                    "type": "string",
                    "description": "What to extract/learn (e.g., 'optimal cycling protocol for NMC811', 'catalyst screening workflow')"
                },
                "synthesis_type": {
                    "type": "string",
                    "enum": ["reference", "trend", "failure", "method"],
                    "description": (
                        "Type of synthesis: 'reference' (default), 'trend', 'failure', or 'method'"
                    )
                }
            },
            required=["plan_ids", "focus"]
        )

        # 13. LIST KNOWLEDGE
        def list_knowledge() -> str:
            """List all active knowledge entries."""
            print(f"  ⚡ Tool: Listing active knowledge...")

            if not self.orch.active_knowledge:
                return json.dumps({
                    "status": "success",
                    "message": "No active knowledge entries.",
                    "entries": []
                })

            entries = []
            for entry in self.orch.active_knowledge:
                entries.append({
                    "id": entry["id"],
                    "focus": entry["focus"],
                    "source_count": len(entry.get("source_plans", entry.get("source_analyses", []))),
                    "findings_count": len(entry.get("key_findings", [])),
                    "timestamp": entry.get("timestamp")
                })

            return json.dumps({
                "status": "success",
                "total_entries": len(entries),
                "entries": entries
            })

        self._register_tool(
            func=list_knowledge,
            name="list_knowledge",
            description="Show all active knowledge entries synthesized from planning iterations.",
            parameters={},
            required=[]
        )

        # 14. CLEAR KNOWLEDGE
        def clear_knowledge(knowledge_id: str = None) -> str:
            """Remove active knowledge entries. If knowledge_id is None, removes all."""
            print(f"  ⚡ Tool: Clearing knowledge...")

            knowledge_dir = self.orch.base_dir / "knowledge"

            if knowledge_id is None:
                count = len(self.orch.active_knowledge)
                self.orch.active_knowledge.clear()
                if knowledge_dir.exists():
                    for f in knowledge_dir.glob("knowledge_*.json"):
                        f.unlink()
                return json.dumps({
                    "status": "success",
                    "message": f"Cleared all {count} knowledge entries."
                })

            for i, entry in enumerate(self.orch.active_knowledge):
                if entry["id"] == knowledge_id:
                    self.orch.active_knowledge.pop(i)
                    knowledge_file = knowledge_dir / f"{knowledge_id}.json"
                    if knowledge_file.exists():
                        knowledge_file.unlink()
                    return json.dumps({
                        "status": "success",
                        "message": f"Removed knowledge entry: {knowledge_id}"
                    })

            return json.dumps({
                "status": "error",
                "message": f"Knowledge ID not found: {knowledge_id}"
            })

        self._register_tool(
            func=clear_knowledge,
            name="clear_knowledge",
            description=(
                "Remove active knowledge entries. Specify a knowledge_id to remove a "
                "specific entry, or omit to clear all knowledge."
            ),
            parameters={
                "knowledge_id": {
                    "type": "string",
                    "description": "ID of knowledge entry to remove (omit to clear all)"
                }
            },
            required=[]
        )

        # 15. GRADUATE TO SKILL
        def graduate_to_skill(knowledge_id: str, skill_name: str, domain: str = "planning") -> str:
            """
            Convert a knowledge entry into a reusable planning skill (.md file).
            The skill is automatically registered for use in subsequent plan generation.
            """
            from .instruct import PLANNING_KNOWLEDGE_TO_SKILL_INSTRUCTIONS

            print(f"  ⚡ Tool: Graduating knowledge '{knowledge_id}' to skill '{skill_name}'...")

            # Find the knowledge entry
            knowledge_entry = None
            for entry in self.orch.active_knowledge:
                if entry.get("id") == knowledge_id:
                    knowledge_entry = entry
                    break

            if knowledge_entry is None:
                return json.dumps({
                    "status": "error",
                    "message": f"Knowledge ID not found: {knowledge_id}"
                })

            # Build knowledge text
            knowledge_text = f"**Focus:** {knowledge_entry.get('focus', '')}\n"
            knowledge_text += f"**Summary:** {knowledge_entry.get('summary', '')}\n"
            knowledge_text += "**Key Findings:**\n"
            for finding in knowledge_entry.get("key_findings", []):
                knowledge_text += f"- {finding}\n"

            # Collect source planning details
            planning_details_parts = []
            source_ids = knowledge_entry.get("source_plans", [])
            plan_history = self.orch.planner.state.get("plan_history", []) if self.orch.planner.state else []

            for pid in source_ids:
                for plan in plan_history:
                    if str(plan.get("iteration")) == str(pid):
                        parts = [f"### Plan Iteration: {pid} (Stage: {plan.get('stage', 'N/A')})"]
                        for exp in plan.get("proposed_experiments", []):
                            parts.append(f"Experiment: {exp.get('experiment_name', '')}")
                            parts.append(f"Hypothesis: {exp.get('hypothesis', '')}")
                            steps = exp.get("experimental_steps", [])
                            if steps:
                                parts.append(f"Steps: {'; '.join(steps[:10])}")
                            parts.append(f"Expected outcome: {exp.get('expected_outcome', '')}")
                        planning_details_parts.append("\n".join(parts))
                        break

            # Also include feedback history if available
            feedback_history = self.orch.planner.state.get("human_feedback_history", []) if self.orch.planner.state else []
            for fb in feedback_history:
                planning_details_parts.append(
                    f"### User Feedback ({fb.get('phase', 'unknown')}):\n{fb.get('feedback', '')}"
                )

            planning_details = "\n\n".join(planning_details_parts) if planning_details_parts else "No source planning details available."

            # Call LLM to generate skill content
            prompt = PLANNING_KNOWLEDGE_TO_SKILL_INSTRUCTIONS.format(
                skill_name=skill_name,
                domain=domain,
                knowledge_text=knowledge_text,
                planning_details=planning_details,
            )

            try:
                response = self.orch.planner.model.generate_content(
                    contents=[prompt],
                    generation_config=None,
                    safety_settings=None,
                )
                skill_content = response.text if hasattr(response, "text") else str(response)
            except Exception as e:
                return json.dumps({"status": "error", "message": f"LLM call failed: {e}"})

            # Save skill file
            skill_dir = self.orch.base_dir / "graduated_skills"
            skill_dir.mkdir(parents=True, exist_ok=True)
            skill_path = skill_dir / f"{skill_name}.md"
            skill_path.write_text(skill_content)

            # Register the skill
            self.orch.register_skill(str(skill_path))

            # Track the link
            self.orch._graduated_skill_sources[skill_name] = [knowledge_id]

            return json.dumps({
                "status": "success",
                "skill_name": skill_name,
                "skill_path": str(skill_path),
                "source_knowledge_id": knowledge_id,
                "note": f"Skill '{skill_name}' has been registered and will be applied to future plan generation."
            })

        self._register_tool(
            func=graduate_to_skill,
            name="graduate_to_skill",
            description=(
                "Convert a knowledge entry into a reusable planning skill (.md file). "
                "The skill is organized into 5 sections (overview, planning, implementation, "
                "interpretation, validation) and automatically registered for use in "
                "subsequent plan generation."
            ),
            parameters={
                "knowledge_id": {
                    "type": "string",
                    "description": "ID of the knowledge entry to graduate"
                },
                "skill_name": {
                    "type": "string",
                    "description": "Name for the new skill (used as filename and reference)"
                },
                "domain": {
                    "type": "string",
                    "description": "Domain area (default: 'planning')"
                }
            },
            required=["knowledge_id", "skill_name"]
        )

        # 16. UPDATE SKILL
        def update_skill(skill_name: str, knowledge_ids: list = None) -> str:
            """
            Update a graduated skill with new knowledge entries.
            Preserves the old version as {name}.prev.md.
            """
            from .instruct import PLANNING_SKILL_UPDATE_INSTRUCTIONS

            print(f"  ⚡ Tool: Updating skill '{skill_name}'...")

            # Find the existing skill file
            skill_dir = self.orch.base_dir / "graduated_skills"
            skill_path = skill_dir / f"{skill_name}.md"
            if not skill_path.exists():
                return json.dumps({
                    "status": "error",
                    "message": f"Graduated skill not found: {skill_name}"
                })

            existing_skill = skill_path.read_text()

            # Determine source knowledge IDs
            tracked_ids = self.orch._graduated_skill_sources.get(skill_name, [])
            if knowledge_ids:
                new_ids = knowledge_ids
            else:
                # Use all knowledge entries with matching focus
                focus_areas = set()
                for kid in tracked_ids:
                    for k in self.orch.active_knowledge:
                        if k.get("id") == kid:
                            focus_areas.add(k.get("focus", "").lower())
                new_ids = [
                    k["id"] for k in self.orch.active_knowledge
                    if k["id"] not in tracked_ids and k.get("focus", "").lower() in focus_areas
                ]

            if not new_ids:
                return json.dumps({
                    "status": "error",
                    "message": "No new knowledge entries found to update the skill with."
                })

            # Collect new knowledge texts
            new_knowledge_parts = []
            for kid in new_ids:
                for k in self.orch.active_knowledge:
                    if k.get("id") == kid:
                        part = f"### {kid}\n**Focus:** {k.get('focus', '')}\n"
                        part += f"**Summary:** {k.get('summary', '')}\n"
                        part += "**Key Findings:**\n"
                        for f in k.get("key_findings", []):
                            part += f"- {f}\n"
                        new_knowledge_parts.append(part)
                        break

            new_knowledge = "\n\n".join(new_knowledge_parts)

            # Call LLM to produce updated skill
            prompt = PLANNING_SKILL_UPDATE_INSTRUCTIONS.format(
                skill_name=skill_name,
                existing_skill=existing_skill,
                new_knowledge=new_knowledge,
            )

            try:
                response = self.orch.planner.model.generate_content(
                    contents=[prompt],
                    generation_config=None,
                    safety_settings=None,
                )
                updated_content = response.text if hasattr(response, "text") else str(response)
            except Exception as e:
                return json.dumps({"status": "error", "message": f"LLM call failed: {e}"})

            # Save previous version
            prev_path = skill_dir / f"{skill_name}.prev.md"
            prev_path.write_text(existing_skill)

            # Write updated skill
            skill_path.write_text(updated_content)

            # Update source tracking
            all_ids = list(set(tracked_ids + new_ids))
            self.orch._graduated_skill_sources[skill_name] = all_ids

            # Re-register the skill
            self.orch.register_skill(str(skill_path))

            return json.dumps({
                "status": "success",
                "skill_name": skill_name,
                "skill_path": str(skill_path),
                "previous_version": str(prev_path),
                "new_knowledge_ids": new_ids,
                "total_source_ids": all_ids,
                "note": f"Skill '{skill_name}' has been updated. Previous version saved as {prev_path.name}."
            })

        self._register_tool(
            func=update_skill,
            name="update_skill",
            description=(
                "Update a graduated skill with new knowledge entries. "
                "Use when new knowledge has been synthesized and a linked skill "
                "should incorporate the new findings. The old version is preserved."
            ),
            parameters={
                "skill_name": {
                    "type": "string",
                    "description": "Name of the graduated skill to update"
                },
                "knowledge_ids": {
                    "type": "array",
                    "items": {"type": "string"},
                    "description": "Specific knowledge IDs to incorporate (omit to auto-detect from matching focus area)"
                }
            },
            required=["skill_name"]
        )
    
    def _register_tool(self, func: Callable, name: str, description: str, 
                      parameters: Dict[str, Any], required: list = None):
        """
        Register a tool in both OpenAI and Gemini formats.
        
        Args:
            func: The Python function to call
            name: Function name
            description: What the function does
            parameters: Dict of parameter definitions
            required: List of required parameter names
        """
        # Add to function map for execution
        self.functions_map[name] = func
        
        # Add to Gemini format (just the function object)
        self.gemini_functions.append(func)
        
        # Build OpenAI schema
        openai_schema = {
            "type": "function",
            "function": {
                "name": name,
                "description": description,
                "parameters": {
                    "type": "object",
                    "properties": parameters,
                    "required": required or []
                }
            }
        }
        self.openai_schemas.append(openai_schema)

    def _update_skill_description(self, custom_skills: dict = None) -> None:
        """Refresh the ``skill`` parameter description in ``generate_initial_plan``.

        Called after a skill is registered at runtime (e.g. ``graduate_to_skill``)
        so newly available skills become visible to the orchestrator LLM. The
        schema dict is mutated in place, so the change propagates to
        ``tools_for_model`` (which is the same ``openai_schemas`` list object).
        """
        new_desc = _build_planning_skill_description(custom_skills)
        for schema in self.openai_schemas:
            fn = schema.get("function", {})
            if fn.get("name") != "generate_initial_plan":
                continue
            skill_prop = fn.get("parameters", {}).get("properties", {}).get("skill")
            if skill_prop is not None:
                skill_prop["description"] = new_desc
            break

    def execute_tool(self, tool_name: str, **kwargs) -> str:
        """
        Execute a tool by name with given arguments.
        
        Args:
            tool_name: Name of the tool to execute
            **kwargs: Arguments to pass to the tool
            
        Returns:
            JSON string with result
        """
        if tool_name not in self.functions_map:
            return json.dumps({
                "status": "error",
                "message": f"Tool '{tool_name}' not found in registry"
            })
        
        try:
            return self.functions_map[tool_name](**kwargs)
        except Exception as e:
            logging.error(f"Tool execution error ({tool_name}): {e}", exc_info=True)
            return json.dumps({
                "status": "error",
                "message": str(e),
                "tool": tool_name
            })


